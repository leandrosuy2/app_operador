from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate, login

from .models import (
    Devedor,
    Empresa,
    Titulo,
    Acordo,
    Parcelamento,
    UserAccessLog,
    MensagemWhatsapp,
    TabelaRemuneracao,
    TabelaRemuneracaoLista,
    EmailEnvio,
    EmailTemplate,
)
from django.apps import AppConfig
from django.core.paginator import Paginator
from datetime import timedelta
import logging
import base64
import binascii
from datetime import date, datetime
from django.urls import reverse
from django.contrib import messages
from .models import Empresa, Parcelamento, FollowUp
from django.db import connection
from django.http import HttpResponseNotFound, JsonResponse, HttpResponse
from dateutil.relativedelta import relativedelta
from django.utils.dateformat import format
from django.utils import translation
from django.utils import timezone
from django.db.models import F, Sum, Q, Min, Subquery, OuterRef
from django.views.decorators.http import require_POST
from core.models import Acordo, TipoDocTitulo, Agendamento, FollowUp, UsersLojistas
from django.utils.timezone import make_aware, now
import re
from .utils import consultar_cnpj_via_scraping
import time
from django.contrib.auth.models import User, Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.hashers import make_password
from django.utils.translation import gettext_lazy as _
from django.utils.translation import activate
from django.views.decorators.csrf import csrf_exempt
import json
from weasyprint import HTML
from django.template.loader import render_to_string
import tempfile
import traceback
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


from django.contrib.auth.decorators import login_required, permission_required
from django.db import connection
from num2words import num2words
import os
from django.conf import settings
from .forms import MensagemWhatsappForm
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation

import uuid
import bcrypt


import logging

logger = logging.getLogger(__name__)


# core/views.py

def format_brl(valor):
    valor = float(valor)
    inteiro, decimal = f"{valor:.2f}".split(".")
    inteiro = inteiro[::-1]
    partes = [inteiro[i:i+3] for i in range(0, len(inteiro), 3)]
    inteiro_ptbr = ".".join(partes)[::-1]
    return f"{inteiro_ptbr},{decimal}"


def home_redirect(request):
    if request.user.is_authenticated:
        return redirect("dashboard")  # Redireciona apenas para usuários autenticados
    return redirect("login")  # Redireciona para login se não estiver autenticado


class CoreConfig(AppConfig):
    default_auto_field = "django.db.models.BigAutoField"
    name = "core"


def format_whatsapp_number(phone):
    """Formata o número de telefone para o padrão do WhatsApp (sem caracteres especiais e com prefixo 55)."""
    if not phone:
        return None
    # Remove caracteres não numéricos
    phone = re.sub(r"\D", "", phone)
    # Adiciona o código do Brasil (55) se o número não começar com ele
    if not phone.startswith("55"):
        phone = f"55{phone}"
    return phone
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import connection
from django.db.models import Q, F
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta
import json, re

from .models import Titulo, Devedor, Parcelamento, Acordo, Agendamento


@login_required
def assumir_devedor(request, titulo_id):
    # (SEU CÓDIGO ORIGINAL, inalterado)
    username = request.user.username
    with connection.cursor() as cursor:
        cursor.execute(
            "UPDATE titulo SET operador = %s WHERE id = %s",
            [username, titulo_id]
        )
    messages.success(request, "Devedor assumido com sucesso!")
    return HttpResponseRedirect(reverse('detalhes_devedor', args=[titulo_id]))


def format_brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


@login_required
def dashboard(request):
    # -------------------------------------------------------
    # Datas (em horário local) para evitar divergências de TZ
    # -------------------------------------------------------
    local_now = timezone.localtime()
    hoje = local_now.date()
    dia_inicio = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
    dia_fim = dia_inicio + timedelta(days=1)

    query = (request.GET.get("query") or "").strip()

    # Usuário logado
    username = (request.user.username or "").strip()
    full_name = (request.user.get_full_name() or "").strip()

    # =========================================================
    # SUGESTÕES "PARA ASSUMIR": até 10/dia por operador, sem repetir nunca
    # (não depende de model; usa tabela core_sugestao_diaria)
    # =========================================================
    with connection.cursor() as c:
        c.execute(
            "SELECT COUNT(*) FROM core_sugestao_diaria WHERE dia=%s AND operador=%s",
            [hoje, username]
        )
        existentes = int(c.fetchone()[0] or 0)
    faltam = max(0, 10 - existentes)

    if faltam > 0:
        # Insere sugestões novas sem duplicar:
        # - Escolhe 1 título representativo por devedor (maior atraso; desempate: maior valor, menor id)
        # - Disponíveis: dias_atraso > 60 e sem operador atribuído
        # - Exclui devedores já sugeridos no DIA (para qualquer operador) e
        #   devedores já sugeridos ao MESMO operador em qualquer dia (garante “nunca repetir”)
        with connection.cursor() as c:
            c.execute(
                """
                INSERT IGNORE INTO core_sugestao_diaria
                    (dia, operador, devedor_id, titulo_id, created_at, assumido)
                SELECT
                    %s AS dia,
                    %s AS operador,
                    t.devedor_id,
                    t.id,
                    NOW(),
                    0
                FROM titulo t
                WHERE
                    t.dias_atraso > 60
                    AND (t.operador IS NULL OR t.operador = '')
                    -- Escolhe o melhor título por devedor (NOT EXISTS algum melhor)
                    AND NOT EXISTS (
                        SELECT 1
                        FROM titulo t2
                        WHERE t2.devedor_id = t.devedor_id
                          AND (t2.operador IS NULL OR t2.operador = '')
                          AND (
                                t2.dias_atraso > t.dias_atraso
                             OR (t2.dias_atraso = t.dias_atraso AND t2.valor > t.valor)
                             OR (t2.dias_atraso = t.dias_atraso AND t2.valor = t.valor AND t2.id < t.id)
                          )
                    )
                    -- Não sugerir hoje se já foi sugerido a qualquer operador
                    AND NOT EXISTS (
                        SELECT 1
                        FROM core_sugestao_diaria s1
                        WHERE s1.dia = %s AND s1.devedor_id = t.devedor_id
                    )
                    -- Nunca repetir este devedor para ESTE operador
                    AND NOT EXISTS (
                        SELECT 1
                        FROM core_sugestao_diaria s2
                        WHERE s2.operador = %s AND s2.devedor_id = t.devedor_id
                    )
                ORDER BY t.dias_atraso DESC, t.valor DESC, t.id ASC
                LIMIT %s
                """,
                [hoje, username, hoje, username, faltam]
            )

    # Busca as 10 sugestões do dia para exibir
    with connection.cursor() as c:
        c.execute(
            """
            SELECT
              s.titulo_id AS id,
              d.nome,
              COALESCE(d.cpf, d.cnpj) AS cpf_cnpj,
              e.nome_fantasia,
              t.dias_atraso,
              t.valor
            FROM core_sugestao_diaria s
            JOIN titulo    t ON t.id = s.titulo_id
            JOIN devedores d ON d.id = s.devedor_id
            LEFT JOIN core_empresa e ON e.id = d.empresa_id
            WHERE s.dia = %s AND s.operador = %s
            ORDER BY s.id
            LIMIT 10
            """,
            [hoje, username]
        )
        sug_rows = c.fetchall()

    devedores_disponiveis_data = [{
        "id": r[0],  # titulo_id (compatível com seu assumir_devedor atual)
        "nome": r[1],
        "cpf_cnpj": r[2],
        "nome_fantasia": r[3] or "",
        "dias_atraso": int(r[4] or 0),
        "valor": format_brl(r[5] or 0),
        "operador": "",
        "status": "pendente",
    } for r in sug_rows]

    # =========================================================
    # Cards (contagens rápidas)
    # =========================================================
    titulos_pendentes = Titulo.objects.filter(Q(statusBaixa=0) | Q(statusBaixa__isnull=True)).count()
    titulos_quitados = Titulo.objects.filter(statusBaixa=2).count()
    titulos_negociados = Titulo.objects.filter(statusBaixa=3).count()
    total_clientes = Devedor.objects.count()

    # =========================================================
    # Totais do dia (Negociados / Quitados)
    # =========================================================
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT SUM(valor) AS total_negociados_hoje
            FROM titulo
            WHERE statusBaixa = 3
              AND created_at LIKE CONCAT(CURDATE(), '%');
            """
        )
        row = cursor.fetchone()
    negociados_hoje = round(Decimal(row[0] or 0), 2)

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT SUM(COALESCE(valorRecebido, 0)) AS total_quitados_hoje
            FROM titulo
            WHERE data_baixa = CURDATE();
            """
        )
        row = cursor.fetchone()
    quitados_hoje = round(Decimal(row[0] or 0), 2)

    # =========================================================
    # Detalhes "Quitados Hoje"
    # =========================================================
    quitados_hoje_detalhes = Titulo.objects.raw(
        """
        SELECT t.id,
               d.nome,
               COALESCE(d.cpf, d.cnpj) AS cpf_cnpj,
               e.nome_fantasia,
               t.data_baixa,
               t.valorRecebido
        FROM titulo t
        INNER JOIN devedores d    ON t.devedor_id = d.id
        INNER JOIN core_empresa e ON d.empresa_id = e.id
        WHERE t.data_baixa = %s
        """,
        [hoje],
    )
    quitados_hoje_detalhes_data = [
        {
            "nome": q.nome,
            "cpf_cnpj": q.cpf_cnpj,
            "nome_fantasia": q.nome_fantasia,
            "data_baixa": q.data_baixa.strftime("%d/%m/%Y") if q.data_baixa else "-",
            "valorRecebido": format_brl(q.valorRecebido or 0),
        }
        for q in quitados_hoje_detalhes
    ]

    # =========================================================
    # Detalhes "Negociados Hoje"
    # =========================================================
    negociados_hoje_detalhes = Titulo.objects.raw(
        """
        SELECT t.id,
               d.nome,
               COALESCE(d.cpf, d.cnpj) AS cpf_cnpj,
               e.nome_fantasia,
               t.created_at,
               t.valor
        FROM titulo t
        INNER JOIN devedores d    ON t.devedor_id = d.id
        INNER JOIN core_empresa e ON d.empresa_id = e.id
        WHERE t.statusBaixa = 3
          AND DATE(t.created_at) = %s
        """,
        [hoje],
    )
    negociados_hoje_detalhes_data = [
        {
            "nome": n.nome,
            "cpf_cnpj": n.cpf_cnpj,
            "nome_fantasia": n.nome_fantasia,
            "data_negociacao": n.created_at.strftime("%d/%m/%Y %H:%M") if n.created_at else "-",
            "valor": format_brl(n.valor or 0),
        }
        for n in negociados_hoje_detalhes
    ]

    # =========================================================
    # Contador "Negociados em Atraso"
    # =========================================================
    negociados_em_atraso_count = Titulo.objects.filter(statusBaixa=3, dataVencimento__lt=hoje).count()

    # =========================================================
    # Parcelamentos atrasados / vencendo hoje
    # =========================================================
    parcelamentos_atrasados = (
        Parcelamento.objects.filter(
            Q(status="Pendente")
            & Q(data_vencimento_parcela__lte=hoje)
            & ~Q(acordo__titulo__ultima_acao=hoje)
        )
        .select_related(
            "acordo",
            "acordo__titulo",
            "acordo__titulo__devedor",
            "acordo__titulo__devedor__empresa",
        )
        .annotate(qtde_prc=F("acordo__qtde_prc"))
    )

    # =========================================================
    # Últimas movimentações / clientes
    # =========================================================
    ultimos_movimentos = (
        Acordo.objects.select_related("devedor", "titulo")
        .order_by("-id")[:10]
        .values("id", "devedor__nome", "titulo_id", "entrada", "data_entrada", "contato")
    )

    ultimos_clientes = Devedor.objects.order_by("-id")[:10].values(
        "id", "nome", "cpf", "cnpj", "created_at", "nome_fantasia"
    )

    # =========================================================
    # Agendamentos do dia (com fallback quando não há do operador)
    # =========================================================
    base_q = Agendamento.objects.filter(
        Q(status__istartswith="P"),
        Q(data_retorno__gte=dia_inicio),
        Q(data_retorno__lt=dia_fim),
    )

    if request.user.is_staff or request.user.is_superuser:
        ag_qs = base_q
    else:
        op_filter = Q(operador__iexact=username)
        if full_name:
            op_filter |= Q(operador__iexact=full_name)
        ag_qs = base_q.filter(op_filter)
        if not ag_qs.exists():
            ag_qs = base_q

    ag_qs = (
        ag_qs.select_related("devedor", "empresa")
        .values(
            "id",
            "devedor__nome",
            "devedor__cpf",
            "devedor__cnpj",
            "empresa__nome_fantasia",
            "telefone",
            "data_retorno",
            "data_abertura",
            "assunto",
            "operador",
            "status",
        )
        .order_by("data_retorno", "id")
    )

    def _fmt_tel(t):
        if not t:
            return ""
        d = re.sub(r"\D", "", str(t))
        return d if d.startswith("55") else f"55{d}"

    agendamentos_hoje = list(ag_qs)
    for a in agendamentos_hoje:
        a["telefone_formatado"] = _fmt_tel(a.get("telefone"))

    # =========================================================
    # Agenda de trabalho do dia - Pendentes (seu SQL original)
    # =========================================================
    if query:
        search_filter = """
            AND (
                devedores.nome LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                devedores.cpf LIKE %s OR
                devedores.cnpj LIKE %s
            )
        """
        params = [f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%"]
    else:
        search_filter = "AND titulo.operador = %s"
        params = [username]

    agenda_pendentes_query = f"""
        SELECT
            titulo.id AS id,
            devedores.nome,
            core_empresa.nome_fantasia AS nome_fantasia_credor,
            devedores.nome_mae,
            titulo.operador,
            devedores.cpf,
            devedores.cnpj,
            devedores.rg,
            devedores.telefone1,
            devedores.razao_social
        FROM
            devedores, titulo, core_empresa
        WHERE
            titulo.devedor_id = devedores.id
            AND devedores.empresa_id = core_empresa.id
            AND (titulo.statusBaixa = 0 OR titulo.statusBaixa IS NULL)
            AND (titulo.ultima_acao IS NULL OR DATE(titulo.ultima_acao) != CURDATE())
            AND core_empresa.status_empresa = 1
            {search_filter}
        GROUP BY
            titulo.id,
            devedores.nome,
            core_empresa.nome_fantasia,
            devedores.nome_mae,
            devedores.cpf,
            devedores.cnpj,
            devedores.rg,
            titulo.juros,
            devedores.telefone1,
            devedores.razao_social,
            titulo.operador
        ORDER BY
            titulo.id DESC
    """
    agenda_pendentes = Titulo.objects.raw(agenda_pendentes_query, params)

    paginator_pendentes = Paginator(agenda_pendentes, 10)
    page_number_pendentes = request.GET.get("page")
    agenda_pendentes_paginated = paginator_pendentes.get_page(page_number_pendentes)

    # =========================================================
    # Negociados em atraso (filtrados pelo operador logado)
    # =========================================================
    negociados_em_atraso_query = """
        SELECT
            MIN(titulo.id) AS id,
            core_empresa.id AS empresa_id,
            devedores.nome AS devedor_nome,
            core_empresa.nome_fantasia AS empresa_nome,
            devedores.nome_mae AS nome_mae,
            titulo.devedor_id AS devedor_id,
            MIN(titulo.dataVencimento) AS data_vencimento,
            SUM(titulo.valor) AS valor_total,
            MAX(titulo.operador) AS operador
        FROM
            titulo
        JOIN devedores    ON titulo.devedor_id = devedores.id
        JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE
            titulo.statusBaixa = 3
            AND titulo.dataVencimento < CURRENT_DATE
            AND core_empresa.status_empresa = 1
            AND titulo.operador = %s
        GROUP BY
            core_empresa.id,
            devedores.nome,
            devedores.nome_mae,
            titulo.devedor_id,
            core_empresa.nome_fantasia
    """
    negociados_em_atraso = Titulo.objects.raw(negociados_em_atraso_query, [username])

    paginator_negociados = Paginator(list(negociados_em_atraso), 10)
    page_number_negociados = request.GET.get("page_negociados")
    negociados_paginated = paginator_negociados.get_page(page_number_negociados)

    # =========================================================
    # Contexto (incluo também JSON para usar nos modais via JS)
    # =========================================================
    context = {
        "devedores_disponiveis_data": devedores_disponiveis_data,

        "titulos_pendentes": titulos_pendentes,
        "titulos_quitados": titulos_quitados,
        "titulos_negociados": titulos_negociados,
        "total_clientes": total_clientes,

        "negociados_em_atraso_count": negociados_em_atraso_count,
        "parcelamentos_atrasados": parcelamentos_atrasados,
        "ultimos_movimentos": ultimos_movimentos,
        "ultimos_clientes": ultimos_clientes,

        "agendamentos_hoje": agendamentos_hoje,
        "agenda_pendentes_paginated": agenda_pendentes_paginated,
        "negociados_paginated": negociados_paginated,

        "query": query,
        "quitados_hoje": quitados_hoje,
        "negociados_hoje": negociados_hoje,

        # listas para render server-side
        "quitados_hoje_detalhes": quitados_hoje_detalhes_data,
        "negociados_hoje_detalhes": negociados_hoje_detalhes_data,

        # strings JSON seguras para JSON.parse no template
        "quitados_hoje_detalhes_json": json.dumps(quitados_hoje_detalhes_data, ensure_ascii=False),
        "negociados_hoje_detalhes_json": json.dumps(negociados_hoje_detalhes_data, ensure_ascii=False),
    }

    return render(request, "dashboard.html", context)



@login_required
def listar_grupos(request):
    grupos = Group.objects.all()
    return render(request, "grupos_listar.html", {"grupos": grupos})


# Criar os grupos e permissões (executar uma vez ou em um script separado)
@login_required
def criar_grupos():
    # Criar ou obter os grupos
    admin_group, _ = Group.objects.get_or_create(name="Admin")
    lojista_group, _ = Group.objects.get_or_create(name="Lojista")
    operador_group, _ = Group.objects.get_or_create(name="Operador")

    print("Grupos criados ou já existentes:")
    print(f" - {admin_group}")
    print(f" - {lojista_group}")
    print(f" - {operador_group}")


@login_required
def editar_grupo(request, grupo_id):
    grupo = Group.objects.get(id=grupo_id)
    todas_permissoes = Permission.objects.all()  # Todas as permissões disponíveis

    # Associa permissões ao grupo (verifica se estão atribuídas)
    permissoes_detalhadas = [
        {
            "id": permissao.id,
            "codename": permissao.codename,
            "traduzido": permissao.name,
            "atribuida": grupo.permissions.filter(
                id=permissao.id
            ).exists(),  # Verifica associação
        }
        for permissao in todas_permissoes
    ]

    if request.method == "POST":
        for permissao in todas_permissoes:
            input_name = f"permissoes_{permissao.id}"
            atribuir = request.POST.get(input_name) == "sim"

            if atribuir and not grupo.permissions.filter(id=permissao.id).exists():
                grupo.permissions.add(permissao)  # Adiciona permissão
            elif not atribuir and grupo.permissions.filter(id=permissao.id).exists():
                grupo.permissions.remove(permissao)  # Remove permissão

        return redirect("listar_grupos")  # Redireciona após salvar

    return render(
        request,
        "grupos_editar.html",
        {"grupo": grupo, "permissoes_detalhadas": permissoes_detalhadas},
    )


import json  # Adicione esta linha
from django.http    import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

@csrf_exempt
@login_required
def consult_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido.'}, status=405)

    try:
        payload = json.loads(request.body)
        ids = payload.get('devedores', [])
        # para cada id, chame sua função de atualização:
        for deu_id in ids:
            # substitua pela sua lógica real:
            atualizar_devedor_por_api(deu_id)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def finalizar_titulo(request, titulo_id):
    if request.method == "POST":
        titulo = get_object_or_404(Titulo, id=titulo_id)
        titulo.ultima_acao = now().date()
        titulo.save()
    return JsonResponse(
        {"status": "success", "message": "Título finalizado com sucesso!"}
    )
    return JsonResponse(
        {"status": "error", "message": "Método não permitido."}, status=405
    )


@csrf_exempt  # Permite AJAX, mas use o CSRF Token corretamente no cabeçalho
def atualizar_permissao(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            group_id = data.get("group_id")
            permission_id = data.get("permission_id")
            action = data.get("action")

            # Certifique-se de que os IDs são válidos
            grupo = Group.objects.get(id=group_id)
            permissao = Permission.objects.get(id=permission_id)

            if action == "sim":
                grupo.permissions.add(permissao)
            elif action == "nao":
                grupo.permissions.remove(permissao)
            else:
                return JsonResponse({"success": False, "error": "Ação inválida."})

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Método inválido."})


def permission_denied_view(request, exception):
    return render(request, "403.html", status=403)


def salvar_permissoes(request, grupo_id):
    grupo = Group.objects.get(id=grupo_id)
    if request.method == "POST":
        permissoes = Permission.objects.all()
        for permissao in permissoes:
            # Obtém valor do formulário
            valor = request.POST.get(f"permissoes_{permissao.id}", "nao")
            if valor == "sim":
                grupo.permissions.add(permissao)  # Adiciona permissão
            else:
                grupo.permissions.remove(permissao)  # Remove permissão
    return redirect("listar_grupos")


def listar_permissoes_ptbr():
    permissoes_por_modelo = {}
    permissoes = Permission.objects.select_related("content_type").all()

    for permissao in permissoes:
        modelo = permissao.content_type.model
        app_label = permissao.content_type.app_label
        nome = permissao.name

        descricao_traduzida = traduzir_permissao(nome)
        permissoes_por_modelo.setdefault(f"{app_label} - {modelo}", []).append(
            descricao_traduzida
        )

    return permissoes_por_modelo


def traduzir_permissao(permissao):
    # Traduções das permissões padrões do Django
    traducao = {
        "Can add": "Pode adicionar",
        "Can change": "Pode editar",
        "Can delete": "Pode excluir",
        "Can view": "Pode visualizar",
        # Adicione outras traduções personalizadas aqui se necessário
    }

    # Busca e substitui o padrão "Can <ação> <modelo>"
    for termo_en, termo_pt in traducao.items():
        if termo_en in permissao:
            return permissao.replace(termo_en, termo_pt)
    return permissao  # Retorna o original se não encontrar tradução


def listar_permissoes_view(request):
    permissoes = listar_permissoes_ptbr()
    return render(request, "listar_permissoes.html", {"permissoes": permissoes})


def adicionar_usuario(request):
    groups = Group.objects.all()
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        password_confirmation = request.POST.get("password_confirmation")
        group_name = request.POST.get("group")

        # Verificar se as senhas são iguais
        if password != password_confirmation:
            messages.error(request, "As senhas não conferem.")
            return render(
                request,
                "usuarios_adicionar.html",
                {
                    "groups": groups,
                    "username": username,
                    "email": email,
                    "group_selected": group_name,
                },
            )

        try:
            # Criar usuário e associar ao grupo
            user = User.objects.create_user(
                username=username, email=email, password=password
            )
            group = Group.objects.get(name=group_name)
            user.groups.add(group)
            # messages.success(request, "Usuário criado com sucesso!")
            return redirect("listar_usuarios")
        except Exception as e:
            messages.error(request, f"Erro ao criar usuário: {e}")

    return render(request, "usuarios_adicionar.html", {"groups": groups})


@login_required
def listar_usuarios(request):
    usuarios = User.objects.all()
    return render(request, "usuarios_listar.html", {"usuarios": usuarios})


@login_required
def editar_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)
    groups = Group.objects.all()

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        password_confirmation = request.POST.get("password_confirmation")
        group_name = request.POST.get("group")

        # Verificar se as senhas são iguais
        if password and password != password_confirmation:
            messages.error(request, "As senhas não conferem.")
            return render(
                request,
                "usuarios_editar.html",
                {
                    "user": user,
                    "groups": groups,
                    "username": username,
                    "email": email,
                    "group_selected": group_name,
                },
            )

        try:
            # Atualizar o usuário
            user.username = username
            user.email = email
            if password:
                user.set_password(password)
            user.save()

            # Atualizar o grupo
            user.groups.clear()
            group = Group.objects.get(name=group_name)
            user.groups.add(group)

            # messages.success(request, "Usuário atualizado com sucesso!")
            return redirect("listar_usuarios")
        except Exception as e:
            messages.error(request, f"Erro ao editar usuário: {e}")

    # Carregar os grupos e dados do usuário no formulário
    return render(
        request,
        "usuarios_editar.html",
        {
            "user": user,
            "groups": groups,
            "group_selected": (
                user.groups.first().name if user.groups.exists() else None
            ),
        },
    )


@login_required
def excluir_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    # messages.success(request, f'Usuário {user.username} excluído com sucesso!')
    return redirect("listar_usuarios")


@login_required
def detalhar_parcela(request, parcela_id):
    parcela = get_object_or_404(Parcelamento, id=parcela_id)
    return render(request, "detalhar_parcela.html", {"parcela": parcela})


# Devedores - Listar


# core/views.py (OPERADOR)
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import connection
from django.shortcuts import render

@login_required
def devedores_listar(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()  # Pendente|Quitado|Negociado

    where = ["e.status_empresa = 1"]
    params = {}

    if q:
        where.append("""(
            LOWER(COALESCE(d.nome, d.nome_fantasia, '')) LIKE %(q)s OR
            d.cpf  LIKE %(q)s OR
            d.cnpj LIKE %(q)s
        )""")
        params["q"] = f"%{q.lower()}%"

    # 🔒 Filtro por status usando o número da coluna
    if status == "Pendente":
        where.append("COALESCE(t.statusBaixa, 0) = 0")
    elif status == "Quitado":
        where.append("t.statusBaixa = 2")
    elif status == "Negociado":
        where.append("t.statusBaixa = 3")

    where_sql = " AND ".join(where)

    # Tradução para exibir no HTML
    case_status = """
        CASE
            WHEN t.statusBaixa = 2 THEN 'Quitado'
            WHEN t.statusBaixa = 3 THEN 'Negociado'
            WHEN t.statusBaixa = 0 OR t.statusBaixa IS NULL THEN 'Pendente'
            ELSE 'Pendente'
        END
    """

    list_sql = f"""
        SELECT
            d.id AS id,
            COALESCE(d.nome, d.nome_fantasia) AS nome,
            d.cpf, d.cnpj,
            COALESCE(e.nome_fantasia, e.razao_social) AS empresa,
            t.id AS titulo_id,
            {case_status} AS status_baixa,
            t.statusBaixa  AS status_baixa_val  -- opcional para depurar
        FROM devedores d
        JOIN core_empresa e ON e.id = d.empresa_id
        JOIN titulo t       ON t.devedor_id = d.id
        WHERE {where_sql}
        ORDER BY d.id DESC, t.id DESC
    """

    with connection.cursor() as cur:
        cur.execute(list_sql, params)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    # Badges (respeita a busca 'q', ignora o filtro de status)
    totals_sql = f"""
        SELECT
            SUM(CASE WHEN COALESCE(t.statusBaixa,0) = 0 THEN 1 ELSE 0 END) AS pendentes,
            SUM(CASE WHEN t.statusBaixa = 3 THEN 1 ELSE 0 END)           AS negociados,
            SUM(CASE WHEN t.statusBaixa = 2 THEN 1 ELSE 0 END)           AS quitados,
            COUNT(*)                                                     AS total
        FROM devedores d
        JOIN core_empresa e ON e.id = d.empresa_id
        JOIN titulo t       ON t.devedor_id = d.id
        WHERE {" AND ".join([w for w in where if not w.startswith("COALESCE(t.statusBaixa") and not w.startswith("t.statusBaixa")])}
    """
    with connection.cursor() as cur:
        cur.execute(totals_sql, params)
        pend, nego, quit_, tot = cur.fetchone() or (0, 0, 0, 0)

    totals = {
        "pendentes": pend or 0,
        "negociados": nego or 0,
        "quitados": quit_ or 0,
        "total": tot or 0,
    }

    page_obj = Paginator(rows, 10).get_page(request.GET.get("page"))

    return render(
        request,
        "devedores_listar.html",
        {"page_obj": page_obj, "query": q, "status": status, "totals": totals},
    )




@login_required
def baixar_modelo_devedor(request):
    # Consulta as core_empresa.e tipos de documentos no banco
    empresas = Empresa.objects.values_list("nome_fantasia", flat=True)
    tipos_doc = TipoDocTitulo.objects.values_list(
        "id", "name"
    )  # Supondo que essa seja a tabela de tipos de documentos

    # Cria um Workbook com duas abas
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Devedores e Títulos"

    # Defina os campos do modelo
    campos = [
        "empresa_nome_fantasia",
        "tipo_pessoa",
        "cpf",
        "cnpj",
        "nome",
        "nome_mae",
        "rg",
        "razao_social",
        "nome_fantasia",
        "nome_socio",
        "telefone",
        "telefone1",
        "telefone2",
        "telefone3",
        "telefone4",
        "telefone5",
        "telefone6",
        "telefone7",
        "telefone8",
        "telefone9",
        "telefone10",
        "observacao",
        "cep",
        "endereco",
        "bairro",
        "uf",
        "cidade",
        "email1",
        # Campos de Título
        "num_titulo",
        "dataEmissao",
        "dataVencimento",
        "valor",
        "tipo_doc_id",
    ]
    ws1.append(campos)

    # Adiciona a aba com as empresas
    ws2 = wb.create_sheet("Empresas")
    for empresa in empresas:
        ws2.append([empresa])

    # Adiciona uma aba com os tipos de documento
    ws3 = wb.create_sheet("TiposDoc")
    for tipo in tipos_doc:
        ws3.append([f"{tipo[0]} - {tipo[1]}"])  # Exibe o ID e o nome do documento

    # Adiciona uma aba para as opções de tipo_pessoa
    ws4 = wb.create_sheet("Opções")
    ws4.append(["Tipo Pessoa"])
    ws4.append(["Física"])
    ws4.append(["Jurídica"])

    # Cria validação de dados (drop-down) para a coluna `empresa_nome_fantasia`
    empresa_dv = DataValidation(
        type="list",
        formula1=f"'Empresas'!$A$1:$A${len(empresas)}",
        allow_blank=False,
        showErrorMessage=True,
    )
    ws1.add_data_validation(empresa_dv)
    for row in range(2, 1002):
        empresa_dv.add(ws1[f"A{row}"])  # Aplica à coluna `empresa_nome_fantasia`

    # Cria validação de dados para `tipo_pessoa`
    tipo_pessoa_dv = DataValidation(
        type="list",
        formula1=f"'Opções'!$A$2:$A$3",
        allow_blank=False,
        showErrorMessage=True,
    )
    ws1.add_data_validation(tipo_pessoa_dv)
    for row in range(2, 1002):
        tipo_pessoa_dv.add(ws1[f"B{row}"])  # Aplica à coluna `tipo_pessoa`

    # Cria validação de dados para `tipo_doc_id`
    tipo_doc_dv = DataValidation(
        type="list",
        formula1=f"'TiposDoc'!$A$1:$A${len(tipos_doc)}",
        allow_blank=False,
        showErrorMessage=True,
    )
    ws1.add_data_validation(tipo_doc_dv)
    for row in range(2, 1002):
        tipo_doc_dv.add(ws1[f"AG{row}"])  # Aplica à coluna `tipo_doc_id` (coluna AG)

    # Gera o arquivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="modelo_importacao_devedor_titulo.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def importar_devedor(request):
    if request.method == "POST" and request.FILES.get("arquivo"):
        try:
            # Lê o arquivo enviado
            arquivo = request.FILES["arquivo"]
            df = pd.read_excel(arquivo)

            for _, row in df.iterrows():
                # Busca a empresa pelo nome_fantasia
                empresa = Empresa.objects.filter(
                    nome_fantasia=row["empresa_nome_fantasia"]
                ).first()
                if not empresa:
                    messages.error(
                        request,
                        f"Empresa não encontrada: {row['empresa_nome_fantasia']}",
                    )
                    continue

                # Formata as datas para o formato ISO (yyyy-mm-dd)
                data_emissao = None
                data_vencimento = None

                try:
                    if not pd.isna(row.get("dataEmissao")):
                        data_emissao = pd.to_datetime(
                            row["dataEmissao"], dayfirst=True
                        ).date()
                except Exception as e:
                    logger.error(
                        f"Erro ao converter dataEmissao: {row.get('dataEmissao')} - {e}"
                    )
                    messages.error(request, f"Erro ao converter dataEmissao: {e}")
                    continue

                try:
                    if not pd.isna(row.get("dataVencimento")):
                        data_vencimento = pd.to_datetime(
                            row["dataVencimento"], dayfirst=True
                        ).date()
                except Exception as e:
                    logger.error(
                        f"Erro ao converter dataVencimento: {row.get('dataVencimento')} - {e}"
                    )
                    messages.error(request, f"Erro ao converter dataVencimento: {e}")
                    continue

                # Cria ou recupera o devedor
                try:
                    devedor, created = Devedor.objects.get_or_create(
                        empresa=empresa,
                        tipo_pessoa=row["tipo_pessoa"],
                        cpf=row.get("cpf", None),
                        cnpj=row.get("cnpj", None),
                        defaults={
                            "nome": row["nome"],
                            "nome_mae": row["nome_mae"],
                            "rg": row["rg"],
                            "razao_social": row["razao_social"],
                            "nome_fantasia": row["nome_fantasia"],
                            "nome_socio": row["nome_socio"],
                            "telefone": row["telefone"],
                            "telefone1": row["telefone1"],
                            "telefone2": row["telefone2"],
                            "telefone3": row["telefone3"],
                            "telefone4": row["telefone4"],
                            "telefone5": row["telefone5"],
                            "telefone6": row["telefone6"],
                            "telefone7": row["telefone7"],
                            "telefone8": row["telefone8"],
                            "telefone9": row["telefone9"],
                            "telefone10": row["telefone10"],
                            "observacao": row["observacao"],
                            "cep": row["cep"],
                            "endereco": row["endereco"],
                            "bairro": row["bairro"],
                            "uf": row["uf"],
                            "cidade": row["cidade"],
                            "email1": row["email1"],
                        },
                    )
                except Exception as e:
                    logger.error(
                        f"Erro ao criar ou recuperar devedor {row.get('nome')}: {e}"
                    )
                    messages.error(request, f"Erro ao criar ou recuperar devedor: {e}")
                    continue

                # Criar o título associado ao devedor
                # Criar o título associado ao devedor
                if not pd.isna(row.get("num_titulo")):
                    tipo_doc_id = row.get("tipo_doc_id")
                    if pd.isna(tipo_doc_id) or not tipo_doc_id:
                        logger.error(
                            f"Tipo de documento ausente ou inválido para o título: {row['num_titulo']}"
                        )
                        messages.error(
                            request,
                            f"Tipo de documento ausente ou inválido para o título: {row['num_titulo']}",
                        )
                        continue

                    try:
                        # Extrai apenas o ID numérico (assumindo formato "2 - Cheque")
                        tipo_doc_id = int(str(tipo_doc_id).split("-")[0].strip())

                        Titulo.objects.create(
                            devedor=devedor,
                            num_titulo=row["num_titulo"],
                            dataEmissao=data_emissao,
                            dataVencimento=data_vencimento,
                            valor=row["valor"],
                            tipo_doc_id=tipo_doc_id,  # Agora é garantido que seja um inteiro
                            acordoComfirmed=0,  # Valor padrão para o campo acordoComfirmed
                        )
                        # logger.info(f"Título criado com sucesso para o devedor: {row['nome']}")
                    except Exception as e:
                        logger.error(
                            f"Erro ao criar título para o devedor {row['nome']} - {e}"
                        )
                        messages.error(
                            request,
                            f"Erro ao criar título para o devedor {row['nome']} - {e}",
                        )
                        continue

            # messages.success(request, "Importação concluída com sucesso.")
        except Exception as e:
            logger.error(f"Erro geral durante a importação: {e}")
            messages.error(request, f"Erro geral durante a importação: {e}")
    return redirect("listar_devedores")


@login_required
def agendamentos_cadastrar(request):
    if request.method == "POST":
        try:
            devedor_id = request.POST.get("devedor_id")
            empresa_id = request.POST.get("empresa_id")
            telefone = request.POST.get("telefone")
            data_abertura = request.POST.get("data_abertura")
            data_retorno = request.POST.get("data_retorno")
            assunto = request.POST.get("assunto")

            if not devedor_id or not empresa_id:
                messages.error(request, "Devedor e Empresa são obrigatórios.")
                return redirect("agendamentos_cadastrar")

            devedor = Devedor.objects.get(id=devedor_id)
            if telefone:
                devedor.telefone = telefone
                devedor.save(update_fields=["telefone"])

            Agendamento.objects.create(
                devedor=devedor,
                empresa_id=empresa_id,
                data_abertura=data_abertura,
                data_retorno=data_retorno,
                assunto=assunto,
                operador=request.user.username,   # <- garante consistência
                status="Pendente",                # <- garante o filtro
                telefone=telefone or None,
            )

            return redirect("agendamentos_listar")

        except Exception as e:
            messages.error(request, f"Erro ao criar agendamento: {e}")

    return render(
        request, "agendamentos_criar.html", {"devedores": Devedor.objects.all()}
    )



# Agendamentos - Listar
@login_required
def agendamentos_listar(request):
    return render(request, "agendamentos_listar.html")


# Página de Login


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Verifica se o usuário pertence ao grupo com group_id=1
            if user.groups.filter(id=1).exists():
                login(request, user)
                return redirect("dashboard")  # Redireciona para o dashboard
            else:
                # Usuário autenticado, mas sem permissão para acessar
                return render(
                    request,
                    "login.html",
                    {"error": "Acesso negado. Você não pertence ao grupo autorizado."},
                )
        else:
            # Credenciais inválidas
            return render(
                request,
                "login.html",
                {"error": "Credenciais inválidas. Verifique seu usuário e senha."},
            )

    return render(request, "login.html")


# core/views.py (OPERADOR)
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import connection
from django.shortcuts import render

@login_required
def listar_devedores(request):
    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()  # Pendente | Quitado | Negociado

    base_cte = """
    WITH status_por_devedor AS (
        SELECT
            t.devedor_id,
            CASE
                WHEN MAX(CASE WHEN (t.statusBaixa=3 OR t.statusBaixaGeral=3) THEN 1 ELSE 0 END)=1 THEN 3
                WHEN MAX(CASE WHEN (t.statusBaixa=2 OR t.statusBaixaGeral=2) THEN 1 ELSE 0 END)=1 THEN 2
                -- se tem título mas nenhum 2/3, fica 0; se nem título tem, será NULL (tratamos depois)
                WHEN COUNT(*) > 0 THEN 0
                ELSE NULL
            END AS status_baixa_num,
            MAX(CASE WHEN (t.statusBaixa=3 OR t.statusBaixaGeral=3) THEN 1 ELSE 0 END) AS any_negociado,
            MAX(CASE WHEN (t.statusBaixa=2 OR t.statusBaixaGeral=2) THEN 1 ELSE 0 END) AS any_quitado,
            COUNT(*) AS qtd_titulos,
            MIN(t.id) AS titulo_id_exemplo
        FROM titulo t
        GROUP BY t.devedor_id
    )
    """

    select_block = """
    SELECT
        COALESCE(spd.status_baixa_num, 0) AS status_baixa_num,
        d.id AS devedor_id,
        d.nome AS devedor_nome,
        d.cpf  AS devedor_cpf,
        d.cnpj AS devedor_cnpj,
        COALESCE(e.nome_fantasia, e.razao_social) AS empresa_nome,
        spd.titulo_id_exemplo,
        d.nome_fantasia,
        d.razao_social,
        COALESCE(spd.qtd_titulos, 0) AS qtd_titulos,
        COALESCE(spd.any_negociado, 0) AS any_negociado,
        COALESCE(spd.any_quitado, 0) AS any_quitado
    FROM devedores d
    JOIN core_empresa e           ON d.empresa_id = e.id
    LEFT JOIN status_por_devedor spd ON spd.devedor_id = d.id
    WHERE e.status_empresa = 1
    """

    where_extra, params = "", []
    if q:
        where_extra += """
          AND (
              d.nome LIKE %s
              OR d.cpf LIKE %s
              OR d.cnpj LIKE %s
              OR d.telefone LIKE %s
              OR e.nome_fantasia LIKE %s
              OR d.nome_fantasia LIKE %s
              OR d.razao_social LIKE %s
          )
        """
        like = f"%{q}%"
        params.extend([like]*7)

    # Totais (sem filtro de status)
    sql_totais = base_cte + select_block + where_extra + " ORDER BY d.id"
    from django.db import connection
    with connection.cursor() as cur:
        cur.execute(sql_totais, params)
        rows_all = cur.fetchall()

    totals = {"pendentes": 0, "negociados": 0, "quitados": 0, "total": 0}
    for r in rows_all:
        st = r[0]
        if st == 3: totals["negociados"] += 1
        elif st == 2: totals["quitados"] += 1
        else: totals["pendentes"] += 1
        totals["total"] += 1

    # Listagem (com filtro de status, se houver)
    sql_list = base_cte + select_block + where_extra
    if status_filter == "Negociado":
        sql_list += " AND COALESCE(spd.any_negociado,0) = 1"
    elif status_filter == "Quitado":
        sql_list += " AND COALESCE(spd.any_negociado,0) = 0 AND COALESCE(spd.any_quitado,0) = 1"
    elif status_filter == "Pendente":
        sql_list += " AND COALESCE(spd.any_negociado,0) = 0 AND COALESCE(spd.any_quitado,0) = 0"

    sql_list += " ORDER BY d.id"

    with connection.cursor() as cur:
        cur.execute(sql_list, params)
        rows = cur.fetchall()

    IDX = {
        "STATUS":0,"ID":1,"NOME":2,"CPF":3,"CNPJ":4,"EMP":5,"TIT":6,"NF":7,"RAZAO":8,"QTD":9
    }
    devedores = [{
        "id": row[IDX["ID"]],
        "titulo_id": row[IDX["TIT"]],                    # pode ser None
        "nome_fantasia": row[IDX["NF"]],
        "razao_social": row[IDX["RAZAO"]],
        "nome": row[IDX["NOME"]],
        "cpf": row[IDX["CPF"]],
        "cnpj": row[IDX["CNPJ"]] or "Não informado",
        "empresa": row[IDX["EMP"]],
        "quantidade_titulos": row[IDX["QTD"]],
        "status_baixa": {0:"Pendente",2:"Quitado",3:"Negociado"}.get(row[IDX["STATUS"]], "Pendente"),
    } for row in rows]

    from django.core.paginator import Paginator
    page_obj = Paginator(devedores, 10).get_page(request.GET.get("page"))

    return render(request,"devedores_listar.html",
        {"page_obj": page_obj, "query": q, "status": status_filter, "totals": totals})


    

# Adicionar Devedor
@login_required
def adicionar_devedor(request):
    empresas = Empresa.objects.all()  # Para preencher o campo empresa
    fields = [
        "cpf",
        "cnpj",
        "nome",
        "nome_mae",
        "rg",
        "razao_social",
        "nome_fantasia",
        "nome_socio",
        "cpf_socio",
        "rg_socio",
        "telefone",
        "telefone1",
        "telefone2",
        "telefone3",
        "telefone4",
        "telefone5",
        "telefone6",
        "telefone7",
        "telefone8",
        "telefone9",
        "telefone10",
        "cep",
        "endereco",
        "bairro",
        "uf",
        "cidade",
        "email1",
        "email2",
    ]

    if request.method == "POST":
        data = request.POST
        empresa = Empresa.objects.get(id=data["empresa_id"])
        devedor_data = {
            field: data.get(field) for field in fields
        }  # Captura todos os campos dinamicamente
        devedor = Devedor.objects.create(
            empresa=empresa,
            tipo_pessoa=data["tipo_pessoa"],
            **devedor_data,  # Descompacta os campos capturados
        )
        # messages.success(request, 'Devedor adicionado com sucesso.')

        # Redireciona para a página de adicionar título para o devedor recém-criado
        return redirect("adicionar_titulo_pg_devedor", devedor_id=devedor.id)

    return render(
        request, "devedores_adicionar.html", {"empresas": empresas, "fields": fields}
    )


@login_required
def adicionar_titulo_pg_devedor(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)
    empresas = Empresa.objects.all()
    tipos_docs = TipoDocTitulo.objects.all()

    if request.method == "POST":
        data = request.POST
        tipo_doc = TipoDocTitulo.objects.get(id=data["tipo_doc_id"])

        # Converte o valor para Decimal no formato correto
        valor_formatado = data["valor"].replace(".", "").replace(",", ".")
        valor_decimal = Decimal(valor_formatado)

        Titulo.objects.create(
            empresa=devedor.empresa,  # Usa a empresa associada ao devedor
            devedor=devedor,
            num_titulo=data["num_titulo"],
            valor=valor_decimal,
            dataVencimento=data["data_vencimento"],
            dataEmissao=data["data_emissao"],  # Adiciona a data de emissão
            tipo_doc=tipo_doc,
            statusBaixa=data.get("status_baixa", 0),
        )
        # messages.success(request, 'Título adicionado com sucesso.')
        return redirect("listar_titulos_por_devedor", devedor_id=devedor.id)

    return render(
        request,
        "titulos_adicionar_pg_devedor.html",
        {
            "devedor": devedor,
            "empresas": empresas,
            "tipos_docs": tipos_docs,
        },
    )


@login_required
def editar_devedor(request, id):
    devedor = get_object_or_404(Devedor, id=id)
    empresas = Empresa.objects.all()

    fields = [
        "id",
        "tipo_pessoa",
        "cpf",
        "cnpj",
        "nome",
        "nome_mae",
        "rg",
        "razao_social",
        "nome_fantasia",
        "nome_socio",
        "cpf_socio",
        "rg_socio",
        "cep",
        "endereco",
        "bairro",
        "uf",
        "cidade",
        "email1",
        "email2",
        "email3",
        "observacao",
        "telefone",
        "telefone1",
        "telefone2",
        "telefone3",
        "telefone4",
        "telefone5",
        "telefone6",
        "telefone7",
        "telefone8",
        "telefone9",
        "telefone10",
        "telefone_valido",
        "telefone1_valido",
        "telefone2_valido",
        "telefone3_valido",
        "telefone4_valido",
        "telefone5_valido",
        "telefone6_valido",
        "telefone7_valido",
        "telefone8_valido",
        "telefone9_valido",
        "telefone10_valido",
    ]

    valido_fields = {
        "telefone_valido",
        "telefone1_valido",
        "telefone2_valido",
        "telefone3_valido",
        "telefone4_valido",
        "telefone5_valido",
        "telefone6_valido",
        "telefone7_valido",
        "telefone8_valido",
        "telefone9_valido",
        "telefone10_valido",
    }

    devedor_data = {field: getattr(devedor, field, "") for field in fields}
    changes = []

    if request.method == "POST":
        empresa_id = request.POST.get("empresa_id")
        tipo_pessoa = request.POST.get("tipo_pessoa")
        valid_options = {"SIM", "NÃO", "NAO VERIFICADO"}

        for field in fields:
            new_value = request.POST.get(field, "").strip()
            old_value = getattr(devedor, field, None)
            if field in valido_fields:
                new_value = (
                    new_value.upper()
                    if new_value in valid_options
                    else "NAO VERIFICADO"
                )

            if new_value != (old_value if old_value is not None else ""):
                setattr(devedor, field, new_value or None)
                changes.append(
                    f"{field.capitalize()} alterado de '{old_value}' para '{new_value or 'vazio'}'."
                )

        if empresa_id:
            devedor.empresa_id = empresa_id
        devedor.tipo_pessoa = tipo_pessoa
        devedor.save()

        return JsonResponse(
            {
                "success": True,
                "message": "Devedor atualizado com sucesso!",
                "changes": changes,
            },
            status=200,
        )

    # Inclui a lista de números no contexto
    return render(
        request,
        "devedores_editar.html",
        {
            "devedor": devedor,
            "devedor_data": devedor_data,
            "empresas": empresas,
            "numeros_telefones": range(1, 11),  # Lista de números de 1 a 10
        },
    )


import json
import logging
import requests
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db import connection

# Dados da API
# Dados da API
API_URL = "https://api.validocadastro.com.br/json/service.aspx"
CHAVE_ACESSO = "TkYNXlaJrdIv3m5HBl21PK2i/r2WPGMP2rSLDdOY5Gdof+rU9r6aNCgKtR4hepe4"


def limpar_cpf_cnpj(cpf_cnpj):
    """Remove caracteres especiais do CPF/CNPJ"""
    return "".join(filter(str.isdigit, cpf_cnpj))


def buscar_dados_api_cliente(request, devedor_id):
    cpf_cnpj = request.GET.get("cpf", "").strip()  # Primeiro tenta pegar o CPF
    if not cpf_cnpj:  # Se CPF estiver vazio, tenta buscar pelo CNPJ
        cpf_cnpj = request.GET.get("cnpj", "").strip()

    if not cpf_cnpj:
        return JsonResponse({"success": False, "message": "CPF ou CNPJ não fornecido."})

    cpf_cnpj = limpar_cpf_cnpj(cpf_cnpj)

    if not cpf_cnpj.isdigit():
        return JsonResponse(
            {
                "success": False,
                "message": "CPF ou CNPJ inválido. Insira apenas números.",
            }
        )

    tipo_pessoa = "F" if len(cpf_cnpj) == 11 else "J"
    data = {
        "CodigoProduto": "332",
        "Versao": "20180521",
        "ChaveAcesso": CHAVE_ACESSO,
        "Parametros": {"TipoPessoa": tipo_pessoa, "CPFCNPJ": cpf_cnpj},
    }

    headers = {"Content-Type": "application/json"}
    response = requests.post(API_URL, json=data, headers=headers)

    if response.status_code != 200:
        return JsonResponse(
            {"success": False, "message": "Erro na comunicação com a API."}
        )

    result = response.json()
    status = (
        result.get("HEADER", {})
        .get("INFORMACOES_RETORNO", {})
        .get("STATUS_RETORNO", {})
        .get("CODIGO")
    )

    if status == "1":
        try:
            with connection.cursor() as cursor:  # 🔹 Usa a conexão do Django
                sql = "INSERT INTO consultas (cpfcnpj, consulta_data) VALUES (%s, %s)"
                consulta_data = response.text
                cursor.execute(sql, [cpf_cnpj, consulta_data])

            return JsonResponse(
                {
                    "success": True,
                    "message": "Consulta concluída com sucesso. Resultados salvos no banco de dados.",
                }
            )

        except Exception as err:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Erro ao conectar ao banco de dados: {str(err)}",
                }
            )

    else:
        erro_descricao = (
            result.get("HEADER", {})
            .get("INFORMACOES_RETORNO", {})
            .get("STATUS_RETORNO", {})
            .get("DESCRICAO", "Erro desconhecido")
        )
        return JsonResponse(
            {"success": False, "message": f"Erro na consulta: {erro_descricao}"}
        )


def normalizar_cpf_cnpj(valor):
    """Remove caracteres especiais do CPF/CNPJ para garantir compatibilidade com o banco."""
    return re.sub(r"\D", "", valor)  # Remove tudo que não for número


@csrf_exempt
def salvar_dados_api_cadastro(request):
    if request.method == "POST":
        cpf_cnpj = request.POST.get("cpfcnpj", "").strip()
        if not cpf_cnpj:
            return JsonResponse(
                {"success": False, "message": "CPF ou CNPJ não fornecido."}
            )

        cpf_cnpj_normalizado = normalizar_cpf_cnpj(cpf_cnpj)

        try:
            with connection.cursor() as cursor:
                # Busca os dados na tabela `consultas` pelo CPF/CNPJ
                cursor.execute(
                    "SELECT consulta_data FROM consultas WHERE cpfcnpj = %s",
                    [cpf_cnpj_normalizado],
                )
                resultado = cursor.fetchone()

                if not resultado:
                    return JsonResponse(
                        {
                            "success": False,
                            "message": "Nenhum dado encontrado para esse CPF/CNPJ.",
                        }
                    )

                data = json.loads(resultado[0])
                cred_cadastral = data.get("CREDCADASTRAL", {})
                dados_receita = cred_cadastral.get("DADOS_RECEITA_FEDERAL", {})

                nome = dados_receita.get("NOME", "").strip()
                nome_mae = dados_receita.get("NOME_MAE", "").strip()

                telefones = []
                for tipo in ["TELEFONE_FIXO", "TELEFONE_CELULAR"]:
                    if tipo in cred_cadastral:
                        for tel in cred_cadastral[tipo].get("TELEFONES", []):
                            ddd = tel.get("DDD", "").strip()
                            numero = tel.get("NUM_TELEFONE", "").strip()
                            if ddd and numero:
                                telefones.append(f"({ddd}) {numero}")

                while len(telefones) < 10:
                    telefones.append(None)

                # Atualiza os dados na tabela `devedores`
                update_query = """
                    UPDATE devedores
                    SET nome_socio = %s,
                        nome_mae = %s,
                        telefone1 = %s,
                        telefone2 = %s,
                        telefone3 = %s,
                        telefone4 = %s,
                        telefone5 = %s,
                        telefone6 = %s,
                        telefone7 = %s,
                        telefone8 = %s,
                        telefone9 = %s,
                        telefone10 = %s
                    WHERE REPLACE(REPLACE(REPLACE(cpf, '.', ''), '-', ''), '/', '') = %s;
                """
                params = (
                    nome,
                    nome_mae,
                    telefones[0],
                    telefones[1],
                    telefones[2],
                    telefones[3],
                    telefones[4],
                    telefones[5],
                    telefones[6],
                    telefones[7],
                    telefones[8],
                    telefones[9],
                    cpf_cnpj_normalizado,
                )

                cursor.execute(update_query, params)
                connection.commit()

                if cursor.rowcount > 0:
                    return JsonResponse(
                        {"success": True, "message": "Dados atualizados com sucesso!"}
                    )
                else:
                    return JsonResponse(
                        {
                            "success": False,
                            "message": "Nenhuma linha foi atualizada. O CPF/CNPJ pode não existir no banco.",
                        }
                    )

        except Exception as e:
            return JsonResponse(
                {"success": False, "message": f"Erro ao salvar dados: {str(e)}"}
            )

    return JsonResponse({"success": False, "message": "Método inválido."})


@login_required
# Excluir Devedor
def excluir_devedor(request, id):
    devedor = get_object_or_404(Devedor, id=id)
    if request.method == "POST":
        devedor.delete()
        # messages.success(request, 'Devedor excluído com sucesso.')
        return redirect("listar_devedores")
    return render(request, "devedores_excluir.html", {"devedor": devedor})


@login_required
def titulos_listar(request):
    # Obtém os parâmetros de busca e filtro
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")

    # Mapeamento de status
    status_map = {
        "Pendente": (0, "NULL"),  # Considera 0 e NULL para pendentes
        "Quitado": 2,
        "Negociado": 3,
    }

    # Construindo a consulta SQL base
    query_sql = """
        SELECT 
            titulo.id, 
            core_empresa.razao_social, 
            titulo.num_titulo, 
            titulo.valor, 
            titulo.dataVencimento AS data_vencimento,
            titulo.data_baixa, 
            titulo.statusBaixa AS status_baixa,
            devedores.nome AS devedor_nome,
            devedores.cpf AS devedor_cpf,
            core_empresa.nome_fantasia,            
            devedores.cnpj
        FROM 
            titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE 1=1 and core_empresa.status_empresa =1
    """

    # Adiciona condição de busca, se aplicável
    query_params = []
    if query:
        query_sql += """
            AND (
                titulo.num_titulo LIKE %s OR
                devedores.nome LIKE %s OR
                devedores.cpf LIKE %s OR
                core_empresa.razao_social LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                devedores.cnpj LIKE %s
            )
        """
        query_params = [f"%{query}%"] * 6

    # Adiciona condição de filtro por status, se aplicável
    if status_filter:
        if status_filter == "Pendente":
            query_sql += " AND (titulo.statusBaixa = 0 OR titulo.statusBaixa IS NULL)"
        else:
            query_sql += " AND titulo.statusBaixa = %s"
            query_params.append(status_map[status_filter])

    # Ordena resultados
    query_sql += " ORDER BY titulo.id DESC"

    # Executa a consulta
    with connection.cursor() as cursor:
        cursor.execute(query_sql, query_params)
        rows = cursor.fetchall()

    # Mapeia os resultados
    titulos = [
        {
            "id": row[0],
            "razao_social": row[1],
            "num_titulo": row[2],
            "valor": row[3],
            "data_vencimento": row[4],
            "data_baixa": row[5],
            "status_baixa": row[6],
            "devedor_nome": row[7],
            "devedor_cpf": row[8],  # Adicionando o CPF
            "nome_fantasia": row[9],
            "cnpj": row[10],
        }
        for row in rows
    ]

    # Configura paginação
    paginator = Paginator(titulos, 30)  # Limita 30 registros por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "titulos_listar.html",
        {"page_obj": page_obj, "query": query, "status": status_filter},
    )


from datetime import date
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.shortcuts import render

@login_required
def listar_titulos_por_devedor(request, devedor_id):
    """
    Lista os títulos associados a um devedor específico (com suporte a comprovante).
    """
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT 
                titulo.id                   AS titulo_id,
                core_empresa.razao_social   AS empresa_razao_social,
                titulo.num_titulo           AS numero_titulo,
                titulo.valor                AS valor_titulo,
                titulo.dataVencimento       AS data_vencimento,
                titulo.data_baixa           AS data_baixa,
                titulo.statusBaixa          AS status_baixa,
                devedores.nome              AS nome_devedor,
                titulo.valorRecebido,
                titulo.forma_pag_Id,
                titulo.idTituloRef,
                titulo.comprovante          AS comprovante_path,
                titulo.juros,
                titulo.dias_atraso
            FROM titulo
            INNER JOIN devedores   ON titulo.devedor_id  = devedores.id
            INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
            WHERE devedores.id = %s
            ORDER BY titulo.id DESC
            """,
            [devedor_id],
        )
        # nomes das colunas na ordem retornada
        cols = [c[0] for c in cursor.description]
        idx = {name: i for i, name in enumerate(cols)}
        rows = cursor.fetchall()

    # Mapa de formas de pagamento
    forma_pagamento_map = {
        0: "Pix",
        1: "Dinheiro",
        2: "Cartão de Débito",
        3: "Cartão de Crédito",
        4: "Cheque",
        5: "Depósito em Conta",
        6: "Pagamento na Loja",
        7: "Boleto Bancário",
        8: "Duplicata",
        9: "Recebimento pelo credor",
    }

    # Listas/Acumuladores
    titulos_principais = []
    titulos_entrada = []
    entrada_ids = set()

    total_quitado = 0
    total_negociado = 0
    total_pendente = 0

    # helper local para montar URL pública do arquivo
    from django.conf import settings
    def _media_url(path):
        return f"{settings.MEDIA_URL}{path}" if path else None

    for row in rows:
        valor           = row[idx['valor_titulo']] or 0
        juros           = row[idx['juros']] or 0
        valor_recebido  = row[idx['valorRecebido']] or 0
        status_baixa    = row[idx['status_baixa']] or 0  # trata None como 0

        titulo_dict = {
            "id":               row[idx['titulo_id']],
            "razao_social":     row[idx['empresa_razao_social']],
            "num_titulo":       row[idx['numero_titulo']],
            "valor":            valor,
            "data_vencimento":  row[idx['data_vencimento']],
            "data_baixa":       row[idx['data_baixa']],
            "status_baixa":     status_baixa,
            "devedor_nome":     row[idx['nome_devedor']],
            "valor_recebido":   valor_recebido,
            "forma_pagamento":  forma_pagamento_map.get(row[idx['forma_pag_Id']], "Não definido"),
            "juros":            juros,
            "dias_atraso":      row[idx['dias_atraso']] or 0,
            "valor_com_juros":  valor + juros,
        }

        # comprovante
        path = row[idx['comprovante_path']]
        titulo_dict["tem_comprovante"] = bool(path)
        titulo_dict["comprovante_url"] = _media_url(path)

        # Totais
        if status_baixa == 2:
            total_quitado += (valor_recebido or 0) + (juros or 0)
        elif status_baixa == 3:
            total_negociado += (valor or 0) + (juros or 0)
        elif status_baixa == 0 or status_baixa is None:
            total_pendente += (valor or 0) + (juros or 0)

        # "Entrada" (idTituloRef nulo e status > 1)
        if row[idx['idTituloRef']] is None and status_baixa > 1:
            titulos_entrada.append(titulo_dict)
            entrada_ids.add(row[idx['titulo_id']])

        titulos_principais.append(titulo_dict)

    return render(
        request,
        "titulos_listar_por_devedor.html",
        {
            "titulos":          titulos_principais,
            "titulos_entrada":  titulos_entrada,
            "entrada_ids":      entrada_ids,
            "devedor_id":       devedor_id,
            "total_quitado":    total_quitado,
            "total_negociado":  total_negociado,
            "total_pendente":   total_pendente,
            "today":            date.today(),
        },
    )



def negociacao_devedor(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)

    # Calcular os totais
    total_quitado = (
        Titulo.objects.filter(devedor=devedor, statusBaixa=2).aggregate(
            Sum("valorRecebido")
        )["valorRecebido__sum"]
        or 0
    )
    total_negociado = (
        Titulo.objects.filter(devedor=devedor, statusBaixa=3).aggregate(Sum("valor"))[
            "valor__sum"
        ]
        or 0
    )
    total_pendente = (
        Titulo.objects.filter(devedor=devedor, statusBaixa__in=[0, None]).aggregate(
            Sum("valor")
        )["valor__sum"]
        or 0
    )

    # Debugging - Printando no console
    print(f"Total Quitado: {total_quitado}")
    print(f"Total Negociado: {total_negociado}")
    print(f"Total Pendente: {total_pendente}")

    # Passando o contexto para o template
    context = {
        "devedor": devedor,
        "total_quitado": total_quitado,
        "total_negociado": total_negociado,
        "total_pendente": total_pendente,
    }
    return render(request, "negociacao_devedor.html", context)


@login_required
def adicionar_titulo(request):
    if request.method == "POST":
        data = request.POST
        empresa = Empresa.objects.get(id=data["empresa_id"])
        devedor = Devedor.objects.get(id=data["devedor_id"])

        Titulo.objects.create(
            empresa=empresa,
            devedor=devedor,
            num_titulo=data["num_titulo"],
            valor=data["valor"],
            data_vencimento=data["data_vencimento"],
            statusBaixa=data.get("status_baixa", 0),
        )
        # messages.success(request, 'Título adicionado com sucesso.')
        return redirect("titulos_listar")

    empresas = Empresa.objects.all()
    devedores = Devedor.objects.all()
    return render(
        request,
        "titulos_adicionar.html",
        {"empresas": empresas, "devedores": devedores},
    )


@login_required
def editar_titulo(request, id):
    titulo = get_object_or_404(Titulo, id=id)

    # Restringir edição para statusBaixa = 0, null ou 3
    if titulo.statusBaixa not in [0, None, 3]:
        messages.error(request, "Este título não pode ser editado.")
        return redirect("detalhes_devedor", titulo_id=titulo.id)

    if request.method == "POST":
        novo_valor = request.POST.get("valor")
        nova_data_vencimento = request.POST.get("dataVencimento")

        try:
            titulo.valor = float(novo_valor)
            titulo.dataVencimento = datetime.strptime(
                nova_data_vencimento, "%Y-%m-%d"
            ).date()
            titulo.save()
            messages.success(request, "Título atualizado com sucesso!")
        except ValueError:
            messages.error(
                request, "Erro ao atualizar o título. Verifique os valores inseridos."
            )

        return redirect("detalhes_devedor", titulo_id=titulo.id)

    return render(request, "editar_titulo.html", {"titulo": titulo})


@login_required
def excluir_titulo(request, id):
    titulo = get_object_or_404(Titulo, id=id)
    if request.method == "POST":
        titulo.delete()
        # messages.success(request, 'Título excluído com sucesso.')
        return redirect("titulos_listar")

    return render(request, "titulos_excluir.html", {"titulo": titulo})


@login_required
def listar_empresas(request):
    query = request.GET.get("q", "")

    if query:
        empresas = Empresa.objects.filter(
            Q(id__icontains=query)
            | Q(razao_social__icontains=query)
            | Q(nome_fantasia__icontains=query)
            | Q(cnpj__icontains=query)
        )
    else:
        empresas = Empresa.objects.all()

    paginator = Paginator(empresas, 10)  # Paginate with 10 items per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request, "empresas_listar.html", {"page_obj": page_obj, "query": query}
    )


def gerar_contrato_lojista(request, id):
    # Obter os detalhes da empresa
    empresa = get_object_or_404(Empresa, id=id)

    hoje = timezone.localtime(timezone.now())
    data_formatada = empresa.created_at.strftime("%d/%m/%Y")

    # Função auxiliar para evitar erro com valores None ou strings
    def valor_extenso(valor):
        try:
            # Converte o valor para float, se for string
            valor_float = float(valor)
            return valor_por_extenso(valor_float)
        except (ValueError, TypeError):
            # Retorna "Zero" caso o valor não seja numérico ou seja None
            return "Zero"

    # Garantir que o valor de adesão seja numérico antes de passar ao contexto
    valor_adesao = (
        empresa.valor_adesao
        if isinstance(empresa.valor_adesao, (int, float))
        else float(str(empresa.valor_adesao).replace(",", "."))
    )

    # Obter dados da tabela de remuneração
    remuneracoes = TabelaRemuneracaoLista.objects.all().order_by("de_dias")

    # Preparar o contexto com os dados da empresa e da tabela de remuneração
    context = {
        "empresa": {
            "nome_fantasia": empresa.nome_fantasia,
            "razao_social": empresa.razao_social,
            "cnpj": empresa.cnpj,
            "valor_adesao": valor_adesao,
            "valor_adesao_extenso": valor_extenso(
                valor_adesao
            ),  # Adicionado o valor de adesão por extenso
            "endereco": empresa.endereco,  # Outros campos necessários
            "bairro": empresa.bairro,
            "cidade": empresa.cidade,
            "uf": empresa.uf,
            "cpf_contato": empresa.cpf_contato,
            "data_formatada": data_formatada,
        },
        "remuneracoes": remuneracoes,  # Inclui dados da tabela de remuneração no contexto
    }

    # Renderizar o template do contrato com o contexto fornecido
    return render(request, "contrato_template_lojista.html", context)


@login_required
def gerar_ficha_lojista(request, id):
    activate("pt-br")
    hoje = timezone.localtime(timezone.now())
    # Busca os detalhes da empresa com base no ID
    empresa = get_object_or_404(Empresa, id=id)

    # Converte o valor de adesão para float corretamente
    try:
        valor_adesao_float = float(
            str(empresa.valor_adesao).replace(".", "").replace(",", ".")
        )  # Garante a conversão correta
    except ValueError:
        # Função auxiliar para evitar erro com valores None ou strings
        def valor_extenso(valor):
            try:
                # Converte o valor para float, se for string
                valor_float = float(valor)
                return valor_por_extenso(valor_float)
            except (ValueError, TypeError):
                # Retorna "Zero" caso o valor não seja numérico ou seja None
                return "Zero"

            # Garantir que o valor de adesão seja numérico antes de passar ao contexto
            valor_adesao = (
                empresa.valor_adesao
                if isinstance(empresa.valor_adesao, (int, float))
                else float(str(empresa.valor_adesao).replace(",", "."))
            )

    data_formatada = empresa.created_at.strftime("%d/%m/%Y")

    # Converte o valor de adesão para float corretamente
    valor_adesao_float = float(str(empresa.valor_adesao).replace(",", "."))
    # Contexto com as informações da ficha cadastral
    context = {
        "empresa": {
            "razao_social": empresa.razao_social,
            "cnpj": empresa.cnpj,
            "nome_fantasia": empresa.nome_fantasia,
            "endereco": empresa.endereco,
            "cidade": empresa.cidade,
            "uf": empresa.uf,
            "cep": empresa.cep,
            "telefone": empresa.telefone,
            "email": empresa.email,
            "valor_adesao": empresa.valor_adesao,
            "contratante_nome": empresa.nome_contato,  # Ajuste conforme os campos do modelo
            "contratante_cpf": empresa.cpf_contato,
            "data_cadastro": data_formatada,
            "valor_adesao_extenso": valor_por_extenso(valor_adesao_float),
        },
        "dados_ficha": {
            "valor_mensal": "R$ isento",
            "site": "sist.nortecheck.com.br",
            "fidelidade": "Os títulos deverão permanecer no sistema de cobrança por no mínimo (30) Trinta dias, para solicitação de cancelamento, após a data de inclusão. A exclusão neste período será considerada como baixa e aplicado o Percentual de Cobrança por Títulos.",
            "observacao": "Os títulos lançados no sistema Nortecheck cobrança, terão uma carência de (3) três dias para solicitação de baixa sem custo, não havendo contato por telefone ou SMS.",
        },
        "cidade_data": "Alvorada-TO, 22 de janeiro de 2025",
    }

    # Renderiza o template com os dados fornecidos
    return render(request, "ficha_template_lojista.html", context)


@login_required
def editar_empresa(request, id):
    empresa = get_object_or_404(Empresa, id=id)
    tabelas = TabelaRemuneracao.objects.all()  # Obtém todas as tabelas de remuneração

    if request.method == "POST":
        empresa.razao_social = request.POST.get("razao_social", empresa.razao_social)
        empresa.nome_fantasia = request.POST.get("nome_fantasia", empresa.nome_fantasia)
        empresa.cnpj = request.POST.get("cnpj", empresa.cnpj)
        empresa.nome_contato = request.POST.get("nome_contato", empresa.nome_contato)
        empresa.cpf_contato = request.POST.get("cpf_contato", empresa.cpf_contato)
        empresa.telefone = request.POST.get("telefone", empresa.telefone)
        empresa.celular = request.POST.get("celular", empresa.celular)
        empresa.whatsapp_financeiro = request.POST.get(
            "whatsapp_financeiro", empresa.whatsapp_financeiro
        )
        empresa.banco = request.POST.get("banco", empresa.banco)
        empresa.nome_favorecido_pix = request.POST.get(
            "nome_favorecido_pix", empresa.nome_favorecido_pix
        )
        empresa.tipo_pix = request.POST.get("tipo_pix", empresa.tipo_pix)
        empresa.cep = request.POST.get("cep", empresa.cep)
        empresa.endereco = request.POST.get("endereco", empresa.endereco)
        empresa.numero = request.POST.get("numero", empresa.numero)
        empresa.bairro = request.POST.get("bairro", empresa.bairro)
        empresa.uf = request.POST.get("uf", empresa.uf)
        empresa.cidade = request.POST.get("cidade", empresa.cidade)
        empresa.email = request.POST.get("email", empresa.email)
        empresa.email_financeiro = request.POST.get(
            "email_financeiro", empresa.email_financeiro
        )
        empresa.operador = request.POST.get("operador", empresa.operador)
        empresa.supervisor = request.POST.get("supervisor", empresa.supervisor)
        empresa.gerente = request.POST.get("gerente", empresa.gerente)
        valor_adesao_raw = request.POST.get("valor_adesao", str(empresa.valor_adesao))
        empresa.valor_adesao = valor_adesao_raw.replace(",", ".")
        plano_id = request.POST.get("plano")

        # Atualiza o plano associado
        if plano_id:
            empresa.plano = TabelaRemuneracao.objects.get(id=plano_id)

        # Verifica se um novo logo foi enviado
        novo_logo = request.FILES.get("logo")
        if novo_logo:
            empresa.logo = novo_logo

        # Atualiza o status da empresa (convertendo string para booleano)
        # Atualiza o status da empresa apenas se o campo for enviado
        status_empresa = request.POST.get("status_empresa", None)
        if status_empresa is not None:
            empresa.status_empresa = status_empresa == "True"

        try:
            empresa.save()
            messages.success(request, "Empresa editada com sucesso.")
            return redirect("listar_empresas")
        except Exception as e:
            messages.error(request, f"Erro ao salvar: {e}")

    return render(
        request, "empresas_editar.html", {"empresa": empresa, "tabelas": tabelas}
    )


@login_required
def alterar_status_empresa(request, id):
    if request.method == "POST":
        empresa = get_object_or_404(Empresa, id=id)

        try:
            # Alternar o status da empresa
            empresa.status_empresa = not empresa.status_empresa
            empresa.save()

            return JsonResponse(
                {"success": True, "status_empresa": empresa.status_empresa}
            )
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método inválido"})


@login_required
def excluir_empresa(request, id):
    empresa = get_object_or_404(Empresa, id=id)
    dependentes = Devedor.objects.filter(empresa=empresa)
    if request.method == "POST":
        dependentes.delete()  # Exclui dependentes
        empresa.delete()  # Exclui empresa
        messages.success(request, "Empresa e dependentes excluídos com sucesso.")
        return redirect("listar_empresas")
    return render(
        request,
        "empresas_excluir.html",
        {
            "empresa": empresa,
            "dependentes": dependentes,
        },
    )


def validar_cnpj(cnpj):
    """
    Valida o formato e a estrutura do CNPJ.
    """
    cnpj = re.sub(r"\D", "", cnpj)  # Remove caracteres não numéricos
    if len(cnpj) != 14:
        return False

    # Validação básica para números sequenciais
    if cnpj in (c * 14 for c in "0123456789"):
        return False

    # Cálculo dos dígitos verificadores
    def calcular_digito(cnpj, peso):
        soma = sum(int(cnpj[i]) * peso[i] for i in range(len(peso)))
        resto = soma % 11
        return "0" if resto < 2 else str(11 - resto)

    primeiro_digito = calcular_digito(cnpj[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    segundo_digito = calcular_digito(cnpj[:13], [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])

    return cnpj[12] == primeiro_digito and cnpj[13] == segundo_digito


import smtplib
from decimal import Decimal, InvalidOperation  # ✅ Importação corrigida
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.contrib import messages
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from .models import Empresa, TabelaRemuneracao, EmailEnvio, EmailTemplate


COPIA_EMAIL = "nortecheck-to@hptmail.com"  # Sempre enviar cópia para este e-mail


def buscar_config_email(tipo_envio):
    """Busca as configurações de e-mail do banco para um tipo específico de envio."""
    return EmailEnvio.objects.filter(tipo_envio=tipo_envio).first()


def buscar_template_email(tipo_envio):
    """Busca o template de e-mail do banco para um tipo específico de envio."""
    return EmailTemplate.objects.filter(tipo_envio=tipo_envio).first()


def substituir_variaveis(template, dados):
    """Substitui variáveis dinâmicas no template de e-mail."""
    for chave, valor in dados.items():
        template = template.replace(f"{{{{{chave}}}}}", str(valor))
    return template


def enviar_email_tipo_envio(tipo_envio, destinatario, dados):
    """Envia um e-mail baseado no tipo de envio e nas informações fornecidas, sempre copiando um e-mail fixo."""
    print(f"🔍 Buscando configurações e template para '{tipo_envio}'...")

    config_email = buscar_config_email(tipo_envio)
    template_email = buscar_template_email(tipo_envio)

    if not config_email or not template_email:
        print(f"❌ Falha: Configuração ou template não encontrado para '{tipo_envio}'")
        return False

    corpo_email = substituir_variaveis(template_email.mensagem, dados)

    msg = MIMEMultipart()
    msg["From"] = config_email.email
    msg["To"] = destinatario
    msg["Cc"] = COPIA_EMAIL  # ✅ Adicionando a cópia
    msg["Subject"] = f"Notificação: {tipo_envio}"

    msg.attach(MIMEText(corpo_email, "plain"))

    destinatarios = [destinatario, COPIA_EMAIL]  # Lista com destinatário e cópia

    try:
        print(
            f"📧 Enviando e-mail para {destinatario} e cópia para {COPIA_EMAIL} via {config_email.servidor_smtp}:{config_email.porta}..."
        )

        server = smtplib.SMTP_SSL(config_email.servidor_smtp, config_email.porta)
        server.login(config_email.email, config_email.senha)
        server.sendmail(config_email.email, destinatarios, msg.as_string())
        server.quit()

        print(
            f"✅ E-mail '{tipo_envio}' enviado com sucesso para {destinatario} e {COPIA_EMAIL}!"
        )
        return True
    except Exception as e:
        print(f"❌ Erro ao enviar e-mail '{tipo_envio}': {e}")
        return False


@login_required
def adicionar_empresa(request):
    tabelas = TabelaRemuneracao.objects.all()

    if request.method == "POST":
        try:
            # Recebe os dados do formulário
            nome_fantasia = request.POST.get("nome_fantasia")
            razao_social = request.POST.get("razao_social")
            nome_contato = request.POST.get("nome_contato")
            cpf_contato = request.POST.get("cpf_contato")
            cnpj = request.POST.get("cnpj")
            telefone = request.POST.get("telefone")
            celular = request.POST.get("celular")
            whatsapp_financeiro = request.POST.get("whatsapp_financeiro")
            email = request.POST.get("email")
            email_financeiro = request.POST.get("email_financeiro")
            cep = request.POST.get("cep")
            endereco = request.POST.get("endereco")
            numero = request.POST.get("numero")
            bairro = request.POST.get("bairro")
            cidade = request.POST.get("cidade")
            uf = request.POST.get("uf")
            ie = request.POST.get("ie")
            plano_id = request.POST.get("plano")
            banco = request.POST.get("banco")
            nome_favorecido_pix = request.POST.get("nome_favorecido_pix")
            tipo_pix = request.POST.get("tipo_pix")
            operador = request.POST.get("operador")
            supervisor = request.POST.get("supervisor")
            gerente = request.POST.get("gerente")

            # Formata e converte o valor de adesão
            valor_adesao_raw = (
                request.POST.get("valor_adesao", "0.0")
                .replace(".", "")
                .replace(",", ".")
            )
            try:
                valor_decimal = Decimal(valor_adesao_raw)
            except InvalidOperation:
                messages.error(
                    request, "O valor de adesão está em um formato inválido."
                )
                return render(request, "empresas_adicionar.html", {"tabelas": tabelas})

            # Validação básica
            if nome_fantasia and razao_social and cnpj and plano_id:
                plano = TabelaRemuneracao.objects.get(
                    id=plano_id
                )  # Busca o plano correspondente
                empresa = Empresa.objects.create(
                    nome_fantasia=nome_fantasia,
                    razao_social=razao_social,
                    cnpj=cnpj,
                    telefone=telefone,
                    celular=celular,
                    whatsapp_financeiro=whatsapp_financeiro,
                    email=email,
                    email_financeiro=email_financeiro,
                    cep=cep,
                    endereco=endereco,
                    numero=numero,
                    bairro=bairro,
                    cidade=cidade,
                    nome_contato=nome_contato,
                    cpf_contato=cpf_contato,
                    uf=uf,
                    ie=ie,
                    valor_adesao=valor_decimal,
                    plano=plano,
                    banco=banco,
                    nome_favorecido_pix=nome_favorecido_pix,
                    tipo_pix=tipo_pix,
                    operador=operador,
                    supervisor=supervisor,
                    gerente=gerente,
                )

                # Enviar e-mail após cadastro
                if email:
                    enviar_email_tipo_envio(
                        "Nova Empresa",
                        email,
                        {
                            "core_empresa.nome_contato": empresa.nome_contato,
                            "core_empresa.nome_fantasia": empresa.nome_fantasia,
                        },
                    )

                messages.success(request, "Empresa adicionada com sucesso.")
                return redirect("listar_empresas")

            else:
                messages.error(
                    request, "Por favor, preencha todos os campos obrigatórios."
                )

        except Exception as e:
            messages.error(request, f"Erro ao adicionar empresa: {e}")
            print(f"❌ Erro ao adicionar empresa: {e}")  # Mostra erro nos logs

    return render(request, "empresas_adicionar.html", {"tabelas": tabelas})


def consultar_cnpj_view(request):
    cnpj = request.GET.get("cnpj", "").strip()
    if not cnpj:
        return JsonResponse({"erro": "CNPJ não fornecido"}, status=400)

    resultado = consultar_cnpj_via_scraping(cnpj)
    return JsonResponse(resultado)


def consultar_com_espera(cnpj):
    time.sleep(5)  # Espera 5 segundos entre as consultas
    return consultar_cnpj_via_scraping(cnpj)


from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render, get_object_or_404

from core.models import Titulo  # ajuste o import conforme seu projeto

@login_required
def realizar_acordo(request, titulo_id):
    print(f"Tentando acessar o título com ID: {titulo_id}")

    titulo = get_object_or_404(Titulo, id=titulo_id, devedor_id__isnull=False)
    print(f"Título encontrado: {titulo}")

    diferenca_dias = (date.today() - titulo.dataVencimento).days if titulo.dataVencimento else 0
    juros_mensais = 0.08
    juros_totais = ((titulo.valor * juros_mensais) * (diferenca_dias / 30)) if diferenca_dias > 0 else 0

    # Atualiza o campo juros no objeto título
    titulo.juros = juros_totais
    titulo.save(update_fields=["juros"])

    # Data de vencimento formatada
    data_vencimento_formatada = titulo.dataVencimento.strftime("%d/%m/%Y") if titulo.dataVencimento else None

    if request.method == "POST":
        data = request.POST
        try:
            entrada = float(data.get("entrada", 0))
            qtde_prc = int(data.get("qtde_prc", 0))

            # total de negociação que o usuário definiu/ajustou
            valor_total_negociacao = float(data.get("valor_total_negociacao", 0))

            # listas vindas do formulário (uma por parcela)
            # OBS: nomes "parcelas_valor[]" e "parcelas_data[]" estão no template
            lista_valores = [v for v in data.getlist("parcelas_valor[]") if str(v).strip() != ""]
            lista_datas   = data.getlist("parcelas_data[]")

            # se não veio lista (fallback), usa o valor_por_parcela único
            if not lista_valores:
                valor_por_parcela_unico = float(data.get("valor_por_parcela", 0))
                if qtde_prc > 0 and valor_por_parcela_unico > 0:
                    lista_valores = [f"{valor_por_parcela_unico:.2f}"] * qtde_prc

            # validações básicas
            if entrada < 0 or qtde_prc <= 0:
                raise ValueError("Entrada deve ser >= 0 e a quantidade de parcelas > 0.")

            if len(lista_valores) != qtde_prc:
                raise ValueError("Quantidade de valores de parcelas não confere com a quantidade informada.")

            # soma das parcelas informadas
            soma_parcelas = sum(float(x or 0) for x in lista_valores)
            esperado = max(valor_total_negociacao - entrada, 0)

            # tolerância de 2 centavos para arredondamentos
            if abs(soma_parcelas - esperado) > 0.02:
                raise ValueError(
                    f"A soma das parcelas (R$ {soma_parcelas:.2f}) precisa bater com "
                    f"o total negociado menos a entrada (R$ {esperado:.2f})."
                )

            # Atualiza o título principal (entrada + metadados)
            titulo.statusBaixa = 3
            titulo.valorRecebido = entrada
            titulo.total_acordo = entrada + soma_parcelas
            titulo.valor_parcela = None
            titulo.qtde_parcelas = qtde_prc
            titulo.forma_pag_Id = None
            titulo.primeiro_vencimento = data.get("venc_primeira_parcela")  # string ISO yyyy-mm-dd
            titulo.juros = juros_totais
            titulo.operador = request.user.username  # salva apenas o nome
            titulo.save()

            # cria as parcelas negociadas
            # usamos as datas calculadas pelo front (enviadas em "parcelas_data[]")
            # se não vierem, calculamos aqui a partir do primeiro vencimento
            venc_primeira_parcela = data.get("venc_primeira_parcela")
            base = datetime.strptime(venc_primeira_parcela, "%Y-%m-%d").date() if venc_primeira_parcela else date.today()

            criar_datas = False
            if not lista_datas or len(lista_datas) != qtde_prc:
                criar_datas = True

            for i in range(qtde_prc):
                valor_i = float(lista_valores[i])
                if criar_datas:
                    data_vencimento_i = base + relativedelta(months=i)
                else:
                    data_vencimento_i = datetime.strptime(lista_datas[i], "%Y-%m-%d").date()

                nova_parcela = Titulo.objects.create(
                    idTituloRef=titulo.id,
                    num_titulo=titulo.num_titulo,
                    tipo_doc_id=titulo.tipo_doc_id,
                    dataEmissao=date.today(),
                    dataVencimento=data_vencimento_i,
                    dataVencimentoReal=data_vencimento_i,
                    dataVencimentoPrimeira=venc_primeira_parcela if i == 0 else None,
                    valor=valor_i,
                    qtde_parcelas=qtde_prc,
                    nPrc=i + 1,               # número da parcela
                    forma_pag_Id=None,
                    statusBaixa=3,             # negociada
                    devedor_id=titulo.devedor_id,
                    operador=request.user.username,
                )
                print(f"Parcela {i+1} criada: ID {nova_parcela.id} — R$ {valor_i:.2f} — {data_vencimento_i}")

            messages.success(request, "Acordo realizado com sucesso!")
            print(f"✅ Acordo realizado com sucesso para {titulo.devedor.nome}")

            # (seu envio de email permanece igual — omitido aqui para encurtar)

            return redirect("listar_titulos_por_devedor", titulo.devedor.id)

        except ValueError as e:
            print(f"Erro nos valores fornecidos: {e}")
            messages.error(request, f"Erro nos valores fornecidos: {e}")
        except Exception as e:
            print(f"Erro inesperado ao criar o acordo: {e}")
            messages.error(request, f"Erro inesperado: {e}")

    valor_total_com_juros = (titulo.valor or 0) + (juros_totais or 0)
    context = {
        "titulo": titulo,
        "juros_totais": juros_totais,
        "diferenca_dias": diferenca_dias,
        "valor_total_com_juros": valor_total_com_juros,
        "data_vencimento_formatada": data_vencimento_formatada,
    }
    return render(request, "realizar_acordo.html", context)



def buscar_email_empresa(core_empresa_id):
    """Busca o e-mail da empresa pelo core_empresa.id, garantindo apenas um resultado."""
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT email, email_financeiro
            FROM core_empresa
            WHERE id = %s
            LIMIT 1
        """,
            [core_empresa_id],
        )

        row = cursor.fetchone()
        if row:
            return (
                row[0] or row[1]
            )  # Retorna o e-mail principal ou, se não existir, o financeiro

    return None  # Se não encontrar nada, retorna None


def buscar_email_empresa(core_empresa_id, devedor_id):
    """Busca o e-mail da empresa pelo `core_empresa.id` ou, se não existir, pelo `devedor.empresa_id`."""
    if core_empresa_id:
        query = "SELECT email, email_financeiro FROM core_empresa WHERE id = %s LIMIT 1"
        params = [core_empresa_id]
    else:
        print(
            f"⚠️ `core_empresa_id` está NULL. Tentando buscar via `devedor_id`: {devedor_id}"
        )
        query = """
            SELECT core_empresa.email, core_empresa.email_financeiro 
            FROM devedores
            INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
            WHERE devedores.id = %s
            LIMIT 1
        """
        params = [devedor_id]

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        row = cursor.fetchone()
        if row:
            return (
                row[0] or row[1]
            )  # Retorna o e-mail principal ou, se não existir, o financeiro

    return None  # Se não encontrar nada, retorna None


@login_required
def quitar_parcela(request, titulo_id):
    titulo = get_object_or_404(Titulo, id=titulo_id)

    if request.method == "POST":
        try:
            valor_recebido = float(request.POST.get("valorRecebido"))
            data_baixa = request.POST.get("dataBaixa")
            forma_pagamento = int(request.POST.get("formaPagamento"))

            # Atualizar o título
            titulo.valorRecebido = valor_recebido
            titulo.data_baixa = data_baixa
            titulo.forma_pag_Id = (
                forma_pagamento  # Salvar a forma de pagamento no banco
            )
            titulo.statusBaixa = 2  # Alterar status para Quitado
            titulo.save()

            # Buscar e-mail da empresa usando `core_empresa.id` ou `devedor.empresa_id`
            email_destinatario = buscar_email_empresa(
                titulo.empresa_id, titulo.devedor.id
            )

            if email_destinatario:
                print(
                    f"🔄 Tentando enviar e-mail de quitação para {email_destinatario}..."
                )

                sucesso_email = enviar_email_tipo_envio(
                    "Quitação Parcela",
                    email_destinatario,
                    {
                        "core_empresa.nome_contato": titulo.devedor.empresa.nome_contato
                        or "Cliente",
                        "core_empresa.nome_fantasia": titulo.devedor.empresa.nome_fantasia,
                        "titulo.id": titulo.id,
                        "titulo.valorRecebido": f"R$ {titulo.valorRecebido:,.2f}",
                        "devedores.nome": titulo.devedor.nome,
                    },
                )

                if sucesso_email:
                    print(
                        f"✅ E-mail de quitação enviado com sucesso para {email_destinatario}!"
                    )
                else:
                    print("⚠️ Falha ao enviar e-mail de quitação.")
            else:
                print(
                    f"⚠️ Nenhum e-mail encontrado para a empresa (ID: {titulo.empresa_id}) ou via devedor (ID: {titulo.devedor.id})."
                )

            messages.success(
                request, f"Parcela {titulo.num_titulo} quitada com sucesso!"
            )

        except Exception as e:
            messages.error(request, f"Erro ao quitar parcela: {e}")
            print(f"❌ Erro ao processar quitação da parcela: {e}")

    return redirect("listar_titulos_por_devedor", titulo.devedor_id)


def default_acordo(request):
    if request.user.is_authenticated:
        acordo = Acordo.objects.first()  # Substitua com sua lógica de seleção
        return {"acordo_id": acordo.id if acordo else None}
    return {}


@login_required
def gerar_pdf(request, titulo_id):
    """
    Gera um PDF detalhado para o título especificado usando ReportLab.
    """
    try:
        # Obtenha os dados do título e acordo
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT 
                    devedores.nome AS devedor_nome,
                    devedores.cpf,
                    devedores.cnpj,
                    titulo.valor AS valor_titulo,
                    titulo.juros,
                    core_empresa.nome_fantasia AS empresa_nome_fantasia,
                    core_empresa.cnpj AS empresa_cnpj,
                    core_acordo.valor_total_negociacao,
                    core_acordo.entrada,
                    core_acordo.qtde_prc,
                    core_acordo.data_entrada,
                    core_acordo.venc_primeira_parcela,
                    core_acordo.contato,
                    core_acordo.id AS acordo_id
                FROM 
                    devedores
                INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
                INNER JOIN titulo ON titulo.devedor_id = devedores.id
                INNER JOIN core_acordo ON core_acordo.titulo_id = titulo.id
                WHERE titulo.id = %s
            """,
                [titulo_id],
            )
            acordo_data = cursor.fetchone()

        if not acordo_data:
            return HttpResponse("Acordo não encontrado.", status=404)

        # Mapear os dados para um dicionário
        acordo = {
            "devedor_nome": acordo_data[0],
            "cpf": acordo_data[1],
            "cnpj": acordo_data[2],
            "valor_titulo": acordo_data[3],
            "juros": acordo_data[4],
            "empresa_nome_fantasia": acordo_data[5],
            "empresa_cnpj": acordo_data[6],
            "valor_total_negociacao": acordo_data[7],
            "entrada": acordo_data[8],
            "qtde_prc": acordo_data[9],
            "data_entrada": acordo_data[10],
            "venc_primeira_parcela": acordo_data[11],
            "contato": acordo_data[12],
            "acordo_id": acordo_data[13],
        }

        # Obter as parcelas do acordo
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT parcela_numero, data_vencimento, valor
                FROM core_parcelamento
                WHERE acordo_id = %s
            """,
                [acordo["acordo_id"]],
            )
            parcelas = cursor.fetchall()

        # Criar um buffer para o PDF
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Adicionar título do PDF
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawCentredString(
            width / 2, height - 40, "ACORDO EXTRAJUDICIAL DE RENEGOCIAÇÃO DE DÍVIDA"
        )

        # Adicionar detalhes do acordo
        pdf.setFont("Helvetica", 10)
        # pdf.drawString(50, height - 100, f"Nome do cliente: {acordo['devedor_nome']}")
        # pdf.drawString(50, height - 120, f"CPF/CNPJ: {acordo['cpf'] or acordo['cnpj']}")
        #  pdf.drawString(50, height - 140, f"Empresa: {acordo['empresa_nome_fantasia']}")
        # pdf.drawString(50, height - 160, f"CNPJ Empresa: {acordo['empresa_cnpj']}")
        # pdf.drawString(50, height - 180, f"Valor da Dívida: R$ {acordo['valor_titulo']:.2f}")
        # pdf.drawString(50, height - 200, f"Juros: R$ {acordo['juros']:.2f}")
        # Adicionar introdução ao contrato
        # pdf.drawString(50, height - 100, "ACORDO EXTRAJUDICIAL DE RENEGOCIAÇÃO DE DÍVIDA:")
        pdf.drawString(
            50,
            height - 70,
            f"Eu, {acordo['devedor_nome']}, portador do CPF/CNPJ {acordo['cpf'] or acordo['cnpj']}, confirmo a",
        )
        pdf.drawString(
            50,
            height - 90,
            f"Renegociação da dívida descrita acima em favor da empresa {acordo['empresa_nome_fantasia']},",
        )
        pdf.drawString(
            50,
            height - 110,
            f"De CNPJ {acordo['empresa_cnpj']}. Firmo este Contrato de Confissão e Renegociação de Dívida.",
        )
        pdf.drawString(
            50,
            height - 130,
            f"Valor Total da Negociação: R$ {acordo['valor_total_negociacao']:.2f}",
        )
        pdf.drawString(50, height - 150, f"Entrada: R$ {acordo['entrada']:.2f}")
        pdf.drawString(
            50, height - 170, f"Quantidade de Parcelas: {acordo['qtde_prc']}"
        )
        data_entrada_formatada = datetime.strptime(
            str(acordo["data_entrada"]), "%Y-%m-%d"
        ).strftime("%d/%m/%Y")
        venc_primeira_parcela_formatada = datetime.strptime(
            str(acordo["venc_primeira_parcela"]), "%Y-%m-%d"
        ).strftime("%d/%m/%Y")

        pdf.drawString(50, height - 190, f"Data da Entrada: {data_entrada_formatada}")
        pdf.drawString(
            50,
            height - 210,
            f"Vencimento da Primeira Parcela: {venc_primeira_parcela_formatada}",
        )
        # pdf.drawString(50, height - 190, f"Data da Entrada: {acordo['data_entrada']}")
        # pdf.drawString(50, height - 210, f"Vencimento da Primeira Parcela: {acordo['venc_primeira_parcela']}")
        pdf.drawString(50, height - 230, f"Contato: {acordo['contato']}")

        # Adicionar tabela de parcelas
        # pdf.drawString(50, height - 440, "Parcelas:")
        pdf.line(50, height - 250, width - 50, height - 250)
        pdf.drawString(50, height - 270, "Parcela")
        pdf.drawString(150, height - 270, "Data de Vencimento")
        pdf.drawString(300, height - 270, "Valor")
        y = height - 290

        for parcela in parcelas:
            pdf.drawString(50, y, str(parcela[0]))
            pdf.drawString(150, y, parcela[1].strftime("%d/%m/%Y"))
            pdf.drawString(300, y, f"R$ {parcela[2]:.2f}")
            y -= 20

        # Assinatura
        pdf.drawString(50, y - 30, "Confirmo a renegociação nos termos acima.")
        pdf.line(70, y - 70, width - 70, y - 70)
        pdf.drawCentredString(width / 2, y - 80, f"{acordo['devedor_nome']}")
        pdf.drawCentredString(width / 2, y - 100, f"Assinatura")

        # Finalizar o PDF
        pdf.showPage()
        pdf.save()

        # Obter o conteúdo do PDF do buffer
        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="acordo_titulo_{titulo_id}.pdf"'
        )
        return response

    except Exception as e:
        # Log do erro para depuração
        print(f"Erro ao gerar PDF: {e}")
        return HttpResponse(f"Erro ao gerar PDF: {str(e)}", status=500)


@login_required
def listar_acordos(request):
    query = request.GET.get("q", "")

    sql_query = """
        SELECT 
            titulo.id AS titulo_id,
            titulo.valorRecebido,
            titulo.data_baixa,
            titulo.qtde_parcelas,
            titulo.total_acordo,
            titulo.dataVencimentoPrimeira,
            devedores.telefone1 AS contato,
            devedores.nome AS devedor_nome,
            core_empresa.nome_fantasia AS empresa_nome,
            devedores.cpf,
            devedores.cnpj,
            titulo.comprovante,
            titulo.contrato
            
        FROM 
            titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE 
            titulo.idTituloRef IS NULL 
            AND (titulo.statusBaixa = 2 OR titulo.statusBaixa = 3) and core_empresa.status_empresa =1
    """

    params = []
    if query:
        sql_query += """
            AND (
                devedores.nome LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                devedores.cpf LIKE %s OR
                devedores.cnpj LIKE %s
            )
        """
        params.extend([f"%{query}%"] * 4)

    sql_query += " ORDER BY titulo.id DESC"

    with connection.cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()

    acordos = [
        {
            "titulo_id": row[0],
            "valorRecebido": row[1],
            "data_baixa": row[2].strftime("%d/%m/%Y") if row[2] else "",
            "qtde_parcelas": row[3],
            "total_acordo": row[4],
            "dataVencimentoPrimeira": row[5].strftime("%d/%m/%Y") if row[5] else "",
            "contato": row[6],
            "devedor_nome": row[7],
            "empresa_nome": row[8],
            "cpf": row[9],
            "cnpj": row[10],
            "contrato": row[12],
        }
        for row in rows
    ]

    # Adiciona parcelas para cada título
    for acordo in acordos:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, valor, dataVencimento, data_baixa, statusBaixa, comprovante, contrato
                FROM titulo
                WHERE idTituloRef = %s
                ORDER BY dataVencimento ASC
            """,
                [acordo["titulo_id"]],
            )
            parcelas = cursor.fetchall()
            acordo["parcelas"] = [
                {
                    "id": parcela[0],
                    "valor": parcela[1],
                    "data_vencimento": (
                        parcela[2].strftime("%d/%m/%Y") if parcela[2] else ""
                    ),
                    "data_baixa": parcela[3].strftime("%d/%m/%Y") if parcela[3] else "",
                    "status": "Quitado" if parcela[4] == 2 else "Pendente",
                    "status_baixa": parcela[4],
                    "comprovante": parcela[5],
                    "contrato": parcela[6],
                }
                for parcela in parcelas
            ]

            # Marcar parcelas sem IDs ou com dados inconsistentes
            for parcela in acordo["parcelas"]:
                if not parcela["id"]:
                    parcela["status"] = "Invalida"

    paginator = Paginator(acordos, 10)  # Mostra 10 acordos por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request, "acordos_listar.html", {"page_obj": page_obj, "query": query}
    )


def valor_por_extenso(valor):
    unidades = [
        "",
        "um",
        "dois",
        "três",
        "quatro",
        "cinco",
        "seis",
        "sete",
        "oito",
        "nove",
    ]
    dezenas = [
        "",
        "dez",
        "vinte",
        "trinta",
        "quarenta",
        "cinquenta",
        "sessenta",
        "setenta",
        "oitenta",
        "noventa",
    ]
    centenas = [
        "",
        "cento",
        "duzentos",
        "trezentos",
        "quatrocentos",
        "quinhentos",
        "seiscentos",
        "setecentos",
        "oitocentos",
        "novecentos",
    ]
    especiais = {
        10: "dez",
        11: "onze",
        12: "doze",
        13: "treze",
        14: "quatorze",
        15: "quinze",
        16: "dezesseis",
        17: "dezessete",
        18: "dezoito",
        19: "dezenove",
    }

    def numero_por_extenso(n):
        if n == 0:
            return "zero"
        elif n < 10:
            return unidades[n]
        elif n < 20:
            return especiais[n]
        elif n < 100:
            dezena, unidade = divmod(n, 10)
            return dezenas[dezena] + (f" e {unidades[unidade]}" if unidade else "")
        elif n < 1000:
            centena, resto = divmod(n, 100)
            if n == 100:
                return "cem"
            return centenas[centena] + (
                f" e {numero_por_extenso(resto)}" if resto else ""
            )
        else:
            milhar, resto = divmod(n, 1000)
            milhar_extenso = (
                f"{numero_por_extenso(milhar)} mil" if milhar > 1 else "mil"
            )
            return milhar_extenso + (f" e {numero_por_extenso(resto)}" if resto else "")

    reais, centavos = divmod(round(valor * 100), 100)
    reais_extenso = (
        f'{numero_por_extenso(reais)} real{"s" if reais > 1 else ""}' if reais else ""
    )
    centavos_extenso = (
        f'{numero_por_extenso(centavos)} centavo{"s" if centavos > 1 else ""}'
        if centavos
        else ""
    )

    if reais and centavos:
        return f"{reais_extenso} e {centavos_extenso}"
    return reais_extenso or centavos_extenso


def gerar_contrato(request, titulo_id):
    # Obter o título principal
    titulo = get_object_or_404(Titulo, id=titulo_id, idTituloRef__isnull=True)

    # Obter o devedor e empresa associados
    devedor = titulo.devedor
    empresa = devedor.empresa

    # Obter as parcelas associadas (títulos filhos)
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, valor, dataVencimento, data_baixa, nPrc
            FROM titulo
            WHERE idTituloRef = %s
            ORDER BY nPrc
        """,
            [titulo.id],
        )
        parcelas = cursor.fetchall()

    # Função auxiliar para evitar erro com valores None
    def valor_extenso(valor):
        return valor_por_extenso(valor) if valor is not None else "Zero"

    # Preparar o contexto com os dados disponíveis
    context = {
        "devedores": {
            "nome": devedor.nome,
            "endereco": devedor.endereco,
            "cep": devedor.cep,
            "cidade": devedor.cidade,
            "uf": devedor.uf,
            "cpf": devedor.cpf,
        },
        "core_empresa": {
            "razao_social": empresa.razao_social,
            "endereco": empresa.endereco,
            "bairro": empresa.bairro,
            "cidade": empresa.cidade,
            "uf": empresa.uf,
            "cnpj": empresa.cnpj,
        },
        "titulo": {
            "valor_total_negociacao": titulo.total_acordo,
            "valor_total_negociacao_extenso": valor_extenso(titulo.total_acordo),
            "entrada": titulo.valorRecebido,
            "data_entrada": titulo.data_baixa,
            "entrada_extenso": valor_extenso(titulo.valorRecebido),
            "valor_por_parcela": titulo.parcelar_valor,
            "valor_por_parcela_extenso": valor_extenso(titulo.parcelar_valor),
            "qtde_prc": titulo.qtde_parcelas,
        },
        "parcelas": [
            {
                "parcela_numero": parcela[4],
                "data_vencimento_parcela": parcela[2],
                "valor": parcela[1],
                "valor_extenso": valor_extenso(parcela[1]),
            }
            for parcela in parcelas
        ],
    }

    # Renderizar o template
    return render(request, "contrato_template.html", context)

# core/views.py
import os, uuid
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import datetime
from .models import Titulo   # REMOVA o import de Anexo

@login_required
def realizar_baixa(request, titulo_id):
    titulo = get_object_or_404(Titulo, id=titulo_id)

    forma_pagamento_map = {0:"Pix",1:"Dinheiro",2:"Cartão de Débito",3:"Cartão de Crédito",
                           4:"Cheque",5:"Depósito em Conta",6:"Pagamento na Loja",
                           7:"Boleto Bancário",8:"Duplicata",9:"Recebimento pelo credor"}

    if request.method == "POST":
        try:
            tipo_baixa = request.POST.get("tipo_baixa")
            forma_pagamento_key = int(request.POST.get("forma_pagamento", 0))

            if tipo_baixa == "Quitação":
                valor_quitacao = float(request.POST.get("valor_quitacao", 0))
                data_pagamento = request.POST.get("data_pagamento")
                Titulo.objects.filter(id=titulo.id).update(
                    data_baixa=data_pagamento,
                    valorRecebido=valor_quitacao,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=2,
                )
                titulo.data_baixa = data_pagamento
                titulo.valorRecebido = valor_quitacao
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = 2
                titulo.save()
            elif tipo_baixa == "Parcela":
                valor_parcela = float(request.POST.get("valor_parcela", 0))
                data_pagamento = request.POST.get("data_pagamento")
                novo_valor = (titulo.valorRecebido or 0) + valor_parcela
                status = 1 if novo_valor < titulo.valor else 2
                Titulo.objects.filter(id=titulo.id).update(
                    valorRecebido=novo_valor,
                    data_baixa=data_pagamento,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=status,
                )
                titulo.valorRecebido = novo_valor
                titulo.data_baixa = data_pagamento
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = status
                titulo.save()
            else:
                messages.error(request, "Tipo de Baixa inválido.")
                return redirect("realizar_baixa", titulo_id=titulo_id)

            # salvar comprovante no FileField do Titulo
            files = request.FILES.getlist("comprovantes")
            if files:
                f = files[0]  # apenas o primeiro; seu modelo não suporta múltiplos
                ext = os.path.splitext(f.name)[1] or ""
                name = f"{uuid.uuid4()}{ext}"
                titulo.comprovante.save(name, f)
                titulo.save(update_fields=["comprovante"])

            return redirect("titulos_listar")

        except Exception as e:
            messages.error(request, f"Erro ao salvar Baixa: {e}")
            return redirect("realizar_baixa", titulo_id=titulo_id)

    diferenca_dias = (datetime.today().date() - titulo.dataVencimento).days
    juros_totais = (titulo.valor * 0.08 * (diferenca_dias / 30)) if diferenca_dias > 0 else 0
    return render(request, "realizar_baixa.html", {
        "titulo": titulo,
        "juros_totais": juros_totais,
        "data_vencimento_formatada": titulo.dataVencimento.strftime("%d/%m/%Y"),
    })

# core/views.py
# core/views.py
import os, uuid, logging
from datetime import datetime
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.timezone import is_naive, make_naive
from .models import Titulo

log = logging.getLogger(__name__)

def _save_comprovante_titulo(titulo, uploaded_file):
    """
    Salva 1 arquivo no campo FileField titulo.comprovante mantendo a extensão.
    Retorna (name, url) após salvar.
    """
    if not uploaded_file:
        return None, None
    ext = os.path.splitext(uploaded_file.name)[1] or ""
    name = f"{uuid.uuid4()}{ext}"
    titulo.comprovante.save(name, uploaded_file)   # FileField do modelo
    titulo.save(update_fields=["comprovante"])
    return titulo.comprovante.name, titulo.comprovante.url

@login_required
def realizar_baixa(request, titulo_id):
    titulo = get_object_or_404(Titulo, id=titulo_id)

    forma_pagamento_map = {
        0:"Pix",1:"Dinheiro",2:"Cartão de Débito",3:"Cartão de Crédito",
        4:"Cheque",5:"Depósito em Conta",6:"Pagamento na Loja",
        7:"Boleto Bancário",8:"Duplicata",9:"Recebimento pelo credor",
    }

    if request.method == "POST":
        try:
            tipo_baixa = request.POST.get("tipo_baixa")
            forma_pagamento_key = int(request.POST.get("forma_pagamento", 0))

            if tipo_baixa == "Quitação":
                valor_quitacao = float(request.POST.get("valor_quitacao", 0))
                data_pagamento = request.POST.get("data_pagamento")
                Titulo.objects.filter(id=titulo.id).update(
                    data_baixa=data_pagamento,
                    valorRecebido=valor_quitacao,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=2,
                )
                titulo.data_baixa = data_pagamento
                titulo.valorRecebido = valor_quitacao
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = 2
                titulo.save()

            elif tipo_baixa == "Parcela":
                valor_parcela = float(request.POST.get("valor_parcela", 0))
                data_pagamento = request.POST.get("data_pagamento")
                novo_valor = (titulo.valorRecebido or 0) + valor_parcela
                status = 1 if novo_valor < titulo.valor else 2
                Titulo.objects.filter(id=titulo.id).update(
                    valorRecebido=novo_valor,
                    data_baixa=data_pagamento,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=status,
                )
                titulo.valorRecebido = novo_valor
                titulo.data_baixa = data_pagamento
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = status
                titulo.save()
            else:
                messages.error(request, "Tipo de Baixa inválido.")
                return redirect("realizar_baixa", titulo_id=titulo_id)

            # salvar comprovante(s) na baixa
            arquivos = request.FILES.getlist("comprovantes")
            if hasattr(titulo, "comprovante") and arquivos:
                # salva só o primeiro no campo FileField do título
                _save_comprovante_titulo(titulo, arquivos[0])

            return redirect("titulos_listar")

        except Exception as e:
            log.exception("Erro ao salvar Baixa")
            messages.error(request, f"Erro ao salvar Baixa: {e}")
            return redirect("realizar_baixa", titulo_id=titulo_id)

    diferenca_dias = (datetime.today().date() - titulo.dataVencimento).days
    juros_totais = (titulo.valor * 0.08 * (diferenca_dias / 30)) if diferenca_dias > 0 else 0

    context = {
        "titulo": titulo,
        "juros_totais": juros_totais,
        "data_vencimento_formatada": titulo.dataVencimento.strftime("%d/%m/%Y"),
    }
    return render(request, "realizar_baixa.html", context)

@login_required
def titulo_anexos_json(request, titulo_id):
    """
    Retorna anexos para o modal:
      - título principal: campo FileField 'comprovante'
      - parcelas: todos Titulo com idTituloRef = titulo_id e que tenham 'comprovante'
    """
    try:
        tit = get_object_or_404(Titulo, id=titulo_id)
        itens = []

        def _dt_fmt_from_fs(file_field):
            try:
                ts = os.path.getmtime(file_field.path)
                dt = datetime.fromtimestamp(ts)
                if not is_naive(dt):
                    dt = make_naive(dt)
                return dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                return ""

        # 1) comprovante do título principal
        if hasattr(tit, "comprovante") and tit.comprovante:
            itens.append({
                "id": tit.id,
                "name": os.path.basename(tit.comprovante.name),
                "url": tit.comprovante.url,
                "date": _dt_fmt_from_fs(tit.comprovante),
                "size": getattr(tit.comprovante, "size", 0) or 0,
                "content_type": "",
            })

        # 2) comprovantes das parcelas vinculadas (se existirem)
        parcelas = Titulo.objects.filter(idTituloRef=tit.id).only("id","comprovante")
        for p in parcelas:
            if hasattr(p, "comprovante") and p.comprovante:
                itens.append({
                    "id": p.id,
                    "name": f"parcela_{os.path.basename(p.comprovante.name)}",
                    "url": p.comprovante.url,
                    "date": _dt_fmt_from_fs(p.comprovante),
                    "size": getattr(p.comprovante, "size", 0) or 0,
                    "content_type": "",
                })

        return JsonResponse({"titulo_id": tit.id, "anexos": itens})
    except Exception:
        log.exception("anexos.json erro")
        return JsonResponse({"error": "falha ao listar anexos"}, status=500)


@login_required
def listar_parcelamentos(request):
    # Obter o termo de pesquisa
    query = request.GET.get("q", "")

    # Construir a consulta SQL com filtro, se aplicável
    sql_query = """
        SELECT 
            core_parcelamento.id, 
            core_parcelamento.parcela_numero, 
            core_parcelamento.valor,
            core_parcelamento.data_vencimento,
            core_parcelamento.data_vencimento_parcela,
            core_parcelamento.status, 
            core_parcelamento.created_at,
            core_parcelamento.acordo_id, 
            core_acordo.entrada, 
            core_acordo.qtde_prc, 
            core_acordo.contato, 
            core_acordo.titulo_id,
            devedores.nome AS devedor_nome,
            core_empresa.nome_fantasia AS empresa_nome_fantasia,
            core_parcelamento.forma_pagamento,
            devedores.cpf,
            devedores.cnpj
        FROM 
            core_parcelamento
        INNER JOIN 
            core_acordo 
        ON 
            core_parcelamento.acordo_id = core_acordo.id
        INNER JOIN 
            titulo
        ON 
            core_acordo.titulo_id = titulo.id
        INNER JOIN 
            devedores
        ON 
            titulo.devedor_id = devedores.id
        INNER JOIN 
            core_empresa
        ON 
            devedores.empresa_id = core_empresa.id
        WHERE 1=1
    """

    # Adicionar filtro baseado no termo de pesquisa
    params = []
    if query:
        sql_query += """
            AND (
                core_parcelamento.parcela_numero LIKE %s OR
                devedores.nome LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                core_acordo.contato LIKE %s OR
                core_acordo.titulo_id LIKE %s
            )
        """
        params = [f"%{query}%"] * 5

    sql_query += " ORDER BY core_parcelamento.parcela_numero ASC"

    # Executar a consulta
    with connection.cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()

    # Mapear os resultados para uma estrutura legível no template
    parcelamentos = [
        {
            "id": row[0],
            "parcela_numero": f"{row[1]}/{row[9]}",  # Formata parcela_numero / qtde_prc
            "valor": row[2],
            "data_vencimento": row[3],
            "data_vencimento_parcela": row[4],
            "status": row[5],
            "created_at": row[6],
            "acordo_id": row[7],
            "entrada": row[8],
            "qtde_prc": row[9],
            "contato": row[10],
            "titulo_id": row[11],
            "devedor_nome": row[12],
            "empresa_nome_fantasia": row[13],
            "agendamento_forma_pagamento": row[14],
            "cpf": row[15],
            "cnpj": row[16],
        }
        for row in rows
    ]

    # Paginação (20 itens por página)
    paginator = Paginator(parcelamentos, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Renderizar o template com os dados
    return render(
        request,
        "parcelamentos_listar.html",  # Nome do template para exibir os parcelamentos
        {"page_obj": page_obj, "query": query},
    )


import base64
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from core.decorators import group_required
from core.models import Titulo, Devedor, Empresa

@login_required
def gerar_recibo(request, titulo_id):
    titulo  = get_object_or_404(Titulo, id=titulo_id, statusBaixa=2)
    devedor = get_object_or_404(Devedor, id=titulo.devedor_id)
    empresa = get_object_or_404(Empresa, id=devedor.empresa_id)  # quem recebe

    forma_pagamento_map = {0:"Pix",1:"Dinheiro",2:"Cartão de Débito",3:"Cartão de Crédito",
                           4:"Cheque",5:"Depósito em Conta",6:"Pagamento na Loja",
                           7:"Boleto Bancário",8:"Duplicata"}
    forma_pagamento = forma_pagamento_map.get(getattr(titulo, "forma_pag_Id", None), "Não definido")

    base_str = f"{titulo.id}:{int(float(titulo.valorRecebido)*100)}:{titulo.data_baixa:%Y%m%d}" if titulo.data_baixa else f"{titulo.id}"
    autenticacao_token = base64.urlsafe_b64encode(base_str.encode()).decode()[:20]

    # >>> URL da logo (segura)
    logo_url = ""
    try:
        if getattr(empresa, "logo", None) and getattr(empresa.logo, "name", ""):
            # absoluta para evitar bloqueio em impressão
            logo_url = request.build_absolute_uri(empresa.logo.url)
    except Exception:
        logo_url = ""

    context = {
        "empresa": empresa,          # objeto inteiro (demais dados)
        "logo_url": logo_url,        # <-- use isto no template

        "devedor": {"nome": devedor.nome or devedor.razao_social or "",
                    "cpf_cnpj": devedor.cpf or devedor.cnpj or ""},
        "parcela": {"numero": getattr(titulo, "nPrc", None),
                    "qtde_total": getattr(titulo, "qtde_parcelas", None),
                    "data_vencimento": getattr(titulo, "dataVencimento", None),
                    "data_pagamento": getattr(titulo, "data_baixa", None),
                    "valor_pago": getattr(titulo, "valorRecebido", 0),
                    "forma_pagamento": forma_pagamento},
        "titulo": titulo,
        "numero_titulo": getattr(titulo, "num_titulo", getattr(titulo, "nDoc", titulo.id)),
        "observacao_acordo": getattr(titulo, "acordo_id", titulo.id),
        "consultor": getattr(empresa, "operador", "") or getattr(empresa, "supervisor", "") or "",
        "autenticacao_token": autenticacao_token,
        "data_autenticacao": (titulo.data_baixa if titulo.data_baixa else None),
    }
    return render(request, "recibo.html", context)



@login_required
@require_POST
def pagar_parcela(request, parcela_id):
    # Obter os dados do formulário
    valor_pago = float(request.POST.get("valor_pago", 0))
    data_baixa = request.POST.get("data_baixa")
    forma_pagamento = request.POST.get(
        "forma_pagamento"
    )  # Captura a forma de pagamento

    # Validar se a parcela existe no banco de dados
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT valor FROM core_parcelamento WHERE id = %s", [parcela_id]
        )
        parcela = cursor.fetchone()

        if not parcela:
            messages.error(request, "Parcela não encontrada.")
            return redirect("listar_parcelamentos")

        valor_original = parcela[0]

        # Atualizar a parcela com os dados fornecidos
        cursor.execute(
            """
            UPDATE core_parcelamento
            SET status = %s, data_baixa = %s, forma_pagamento = %s
            WHERE id = %s
        """,
            ["Quitado", data_baixa, forma_pagamento, parcela_id],
        )

    # Exibir mensagem de sucesso e redirecionar
    # messages.success(request, f"Parcela {parcela_id} atualizada com sucesso.")
    return redirect("listar_parcelamentos")


from django.db.models import F
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Agendamento


@login_required
def listar_agendamentos(request):
    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")  # Filtro de status
    page_number = request.GET.get("page", 1)  # Número da página atual

    # Obter todos os agendamentos apenas do usuário logado
    agendamentos = Agendamento.objects.select_related("devedor", "empresa").filter(
        operador=request.user.username  # 🔹 Filtro para mostrar apenas agendamentos do usuário logado
    )

    # Aplicar filtro de busca, se houver
    if query:
        agendamentos = agendamentos.filter(devedor__nome__icontains=query)

    # Aplicar filtro de status, se houver
    if status_filter:
        agendamentos = agendamentos.filter(status=status_filter)

    # Ordenar por status com prioridade e depois por data de retorno decrescente
    agendamentos = agendamentos.annotate(status_priority=F("status")).order_by(
        F("status_priority").desc(nulls_last=True), "-data_retorno"
    )

    # Paginação: 10 itens por página
    paginator = Paginator(agendamentos, 10)
    agendamentos_paginados = paginator.get_page(page_number)

    # Renderizar o template
    return render(
        request,
        "agendamentos_listar.html",
        {
            "agendamentos": agendamentos_paginados,
            "query": query,
            "status_filter": status_filter,
        },
    )


@login_required
def finalizar_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    if request.method == "POST":
        agendamento.status = "Finalizado"
        agendamento.save()
        # messages.success(request, 'Agendamento finalizado com sucesso.')
    return redirect("listar_agendamentos")


@login_required
def criar_agendamento(request):
    """
    View para criar um novo agendamento.
    """
    # Recupera os devedores para o template
    devedores = Devedor.objects.select_related("empresa").all()
    devedores_com_empresas = [
        {
            "id": devedor.id,
            "nome": devedor.nome,
            "empresa_id": devedor.empresa.id if devedor.empresa else None,
            "nome_fantasia": devedor.empresa.nome_fantasia if devedor.empresa else "",
            "telefone": devedor.telefone,  # Telefone do devedor
        }
        for devedor in devedores
    ]

    if request.method == "POST":
        try:
            # Captura os dados do formulário
            devedor_id = request.POST["devedor_id"]
            telefone = request.POST["telefone"]
            data_abertura = make_aware(
                datetime.strptime(request.POST["data_abertura"], "%Y-%m-%dT%H:%M")
            )
            data_retorno = make_aware(
                datetime.strptime(request.POST["data_retorno"], "%Y-%m-%dT%H:%M")
            )
            operador = request.POST.get("operador", "")
            assunto = request.POST.get("assunto", "")  # Captura o assunto

            # Recupera o devedor e a empresa associada
            devedor = get_object_or_404(Devedor, id=devedor_id)
            empresa = devedor.empresa

            # Cria o agendamento
            Agendamento.objects.create(
                devedor=devedor,
                empresa=empresa,
                telefone=telefone,
                data_abertura=data_abertura,
                data_retorno=data_retorno,
                operador=operador,
                assunto=assunto,  # Atribui o assunto aqui
                status="Pendente",  # Status inicial
            )

            # Exibe mensagem de sucesso
            # messages.success(request, "Agendamento criado com sucesso!")
            # print("Agendamento criado com sucesso!")  # Log para depuração

            # Redireciona para a lista de agendamentos
            return redirect("listar_agendamentos")

        except Exception as e:
            # Em caso de erro
            messages.error(request, f"Erro ao criar agendamento: {e}")
            print(f"Erro ao criar agendamento: {e}")  # Log para depuração

    return render(
        request, "agendamentos_criar.html", {"devedores": devedores_com_empresas}
    )

# --- anexar comprovante (AJAX) ---
import os, uuid
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from .models import Titulo

@login_required
def anexar_comprovante_titulo(request, titulo_id):
    titulo = get_object_or_404(Titulo, id=titulo_id)

    if request.method == "POST" and request.FILES.get("comprovante"):
        f = request.FILES["comprovante"]
        ext = os.path.splitext(f.name)[1].lower()
        name = f"{uuid.uuid4()}{ext}"

        titulo.comprovante.save(name, f)
        titulo.save(update_fields=["comprovante"])

        # resposta AJAX
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({
                "ok": True,
                "id": titulo.id,
                "url": titulo.comprovante.url,
                "name": titulo.comprovante.name,
            })

        return redirect("titulos_listar_por_devedor", devedor_id=titulo.devedor_id)

    return JsonResponse({"ok": False, "error": "Nenhum arquivo enviado"}, status=400)

@login_required
def anexar_comprovante(request, parcela_id):
    # Obtenha a parcela diretamente da tabela Titulo
    parcela = get_object_or_404(Titulo, id=parcela_id, idTituloRef__isnull=False)

    if request.method == "POST" and "comprovante" in request.FILES:
        comprovante = request.FILES["comprovante"]

        # Gera um nome de arquivo único com UUID e preserva a extensão original
        extension = os.path.splitext(comprovante.name)[1]  # Obtém a extensão do arquivo
        unique_filename = (
            f"{uuid.uuid4()}{extension}"  # Nome único com extensão original
        )

        # Salva o arquivo com o nome único no campo 'comprovante' da parcela
        parcela.comprovante.save(unique_filename, comprovante)
        parcela.save()  # Salva as mudanças no banco de dados

        # messages.success(request, "Comprovante anexado com sucesso!")
        return redirect(
            "acordos_listar"
        )  # Redireciona de volta à lista de parcelamentos

    messages.error(request, "Falha ao anexar o comprovante. Tente novamente.")
    return redirect("acordos_listar")


def baixar_comprovante(request, titulo_id):
    titulo = Titulo.objects.get(id=titulo_id)
    if titulo.comprovante:
        comprovante_path = (
            titulo.comprovante.path
        )  # Usando o atributo .path que já considera o MEDIA_ROOT

        with open(comprovante_path, "rb") as f:
            response = HttpResponse(f.read(), content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="{os.path.basename(comprovante_path)}"'
            )
            return response
    else:
        return HttpResponse("Nenhum comprovante disponível.", status=404)


import re
from decimal import Decimal, ROUND_HALF_UP

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Value
from django.db.models.functions import Coalesce, Cast
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .models import Titulo, Agendamento, FollowUp
from .utils import consultar_obito  # se não usar, pode remover

dec_field = DecimalField(max_digits=12, decimal_places=2)
zero = Value(0, output_field=dec_field)

def _format_brl(valor):
    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"

@login_required
def detalhes_devedor(request, titulo_id):
    """
    Abre detalhes usando o ID do TÍTULO.
    Calcula juros em Decimal, soma valor+juros nas vencidas, e evita writes desnecessários.
    """
    # ----- Título base (com devedor + empresa via devedor)
    titulo = get_object_or_404(
        Titulo.objects.select_related("devedor", "devedor__empresa"),
        id=titulo_id
    )
    devedor = titulo.devedor
    if not devedor:
        messages.error(request, "Devedor associado a este título não foi encontrado.")
        return redirect("lista_titulos")

    hoje = timezone.localtime().date()

    # ----- Coleções do devedor
    titulos_qs = (
        Titulo.objects
        .filter(devedor=devedor)
        .select_related("devedor", "devedor__empresa")
        .order_by("-id")
    )
    titulos_entrada = titulos_qs.filter(idTituloRef__isnull=True)
    titulos_associados = titulos_qs.filter(idTituloRef__isnull=False)

    # ----- Juros / dias atraso (cálculo em memória; persiste só se vencido e mudou)
    JUROS_MENSAL = Decimal("0.08")  # 8%/mês

    def calc_juros(t):
        if t.dataVencimento and t.statusBaixa != 2 and t.dataVencimento < hoje:
            dias = (hoje - t.dataVencimento).days
            base = Decimal(t.valor or 0)
            frac = (Decimal(dias) / Decimal(30))  # aprox. meses
            juros_calc = (base * JUROS_MENSAL * frac).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            return juros_calc, dias
        return Decimal("0.00"), 0

    titulos = list(titulos_qs)  # vamos iterar mais de uma vez
    for t in titulos:
        juros_calc, dias_calc = calc_juros(t)
        t._juros_calc = juros_calc
        t._dias_calc = dias_calc

        # persiste apenas se há atraso (não quitado) e mudou
        if juros_calc > 0:
            atual_j = Decimal(getattr(t, "juros", 0) or 0)
            atual_d = int(getattr(t, "dias_atraso", 0) or 0)
            if atual_j != juros_calc or atual_d != dias_calc:
                try:
                    t.juros = juros_calc
                    t.dias_atraso = dias_calc
                    t.save(update_fields=["juros", "dias_atraso"])
                except Exception:
                    # não bloqueia a página por erro de persistência
                    pass

    # ----- Totais por status (somente valor do título, como no seu admin)
    total_quitado = titulos_qs.filter(statusBaixa=2).aggregate(
        total=Coalesce(Sum(Cast(F("valor"), dec_field)), zero)
    )["total"] or 0
    total_negociado = titulos_qs.filter(statusBaixa=3).aggregate(
        total=Coalesce(Sum(Cast(F("valor"), dec_field)), zero)
    )["total"] or 0
    total_pendente_valor = titulos_qs.filter(statusBaixa=0).aggregate(
        total=Coalesce(Sum(Cast(F("valor"), dec_field)), zero)
    )["total"] or 0

    # ----- Helper valor+juros (usa juros persistido OU calculado em memória)
    def valor_com_juros(t):
        v = Decimal(t.valor or 0)
        j = Decimal((t.juros if t.juros is not None else t._juros_calc) or 0)
        return (v + j).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    # ----- Conjuntos
    nao_quitados = [t for t in titulos if t.statusBaixa != 2]
    vencidas = [t for t in nao_quitados if t.dataVencimento and t.dataVencimento < hoje]
    a_vencer = [t for t in nao_quitados if not t.dataVencimento or t.dataVencimento >= hoje]
    pendentes = [t for t in titulos if t.statusBaixa == 0]

    # ----- Saldos/valores com juros onde faz sentido
    saldo_pendente = sum((valor_com_juros(t) for t in pendentes), Decimal("0.00"))
    total_vencidas = sum((valor_com_juros(t) for t in vencidas), Decimal("0.00"))
    total_quebra = sum((valor_com_juros(t) for t in nao_quitados), Decimal("0.00"))

    # ----- Extras
    forma_pagamento_map = {
        0: "Pix", 1: "Dinheiro", 2: "Cartão de Débito", 3: "Cartão de Crédito",
        4: "Cheque", 5: "Depósito em Conta", 6: "Pagamento na Loja",
        7: "Boleto Bancário", 8: "Duplicata", 9: "Recebimento pelo credor",
    }
    empresa = getattr(devedor, "empresa", None)
    agendamentos = Agendamento.objects.filter(devedor=devedor)
    follow_ups = FollowUp.objects.filter(devedor=devedor).order_by("-created_at")

    # ----- Máscara doc
    def mascarar_documento(documento):
        doc = re.sub(r"\D", "", (documento or ""))
        if len(doc) == 11:
            return f"{doc[:3]}.{doc[3:6]}.xxx.xx"
        if len(doc) == 14:
            return f"{doc[:2]}.{doc[2:5]}.{doc[5:6]}xxx-xx"
        return "N/A"

    cpf_cnpj = devedor.cpf or devedor.cnpj
    cpf_cnpj_mascarado = mascarar_documento(cpf_cnpj) if cpf_cnpj else "N/A"
    nome_consultor = request.user.get_full_name() or request.user.username
    nome_credor = (getattr(empresa, "nome_fantasia", None) or "NomeCredor")

    # ----- Mensagens Whats (vencidas/quebra com valor+juros)
    qtde_vencidas = len(vencidas)
    lista_vencidas = ", ".join(
        [t.dataVencimento.strftime("%d/%m/%Y") for t in vencidas if t.dataVencimento]
    )

    tpl_vencidas = (
        "Olá %Nome%\n"
        "CPF: %CpfCnpjMascarado%\n\n"
        "Me chamo %NomeConsultor% e tenho uma mensagem importante para você.\n\n"
        "Lamentamos que não tenha pago a (s) parcela (s) referente ao acordo firmado junto a empresa %NomeCredor%.\n\n\n"
        "Parcelas a serem acionadas por meios jurídicos em 5 (cinco) dias úteis.\n\n"
        "Parcelas em aberto: %QtdeParcelas%\n"
        "Vencimentos: %ListaVencimentosParcelas%\n\n"
        "Valor total em aberto: R$ %ValorTotalParcelas%\n\n"
        "Solicitamos sua imediata atenção, hoje seu acordo esta sendo encaminhado para negativação junto aso órgãos de proteção ao credito SPC/SERASA/BOA VISTA e em seguida para analise e acionamento juridico junto a comarca de sua cidade\n\n"
        "PROTOCOLO JUIDICO 160120\n\n"
        "Converse com Negociar Cobranças no WhatsApp: wa.me/5591991600118\n"
        "®NEGOCIAR COBRANÇA\n"
        "CNPJ: 12.855.602/0001-74"
    )
    tpl_a_vencer = (
        "Olá %Nome%\n\n"
        "ME chamo %NomeConsultor%, gostaria de lembrar sobre a (as ) parcela (s) a vencer referente ao acordo firmado junto a empresa %NomeCredor%.\n\n"
        "solicito que ao efetuar o pagamento da mesma, que encaminhe o comprovante nesse meu contato ou no contato direto do nosso financeiro 91991600118\n"
        "Fico no aguardo do envio do comprovante\n\n"
        "NOSSO CANAL DE ATENDIMENTO\n"
        "(91) 99160-0118\n"
        "®NEGOCIAR COBRANÇA\n"
        "Att:.\n"
        "%NomeConsultor%"
    )
    tpl_padrao = (
        "NOTIFICAÇÃO EXTRAJUDICIAL\n\n"
        "Olá, me Chamo %NomeConsultor%.\n\n"
        "Nesse contato eu falo com %Nome%, portado do documento %CpfCnpjMascarado% ?\n\n"
        "Digite - 01 - Para SIM\n"
        "Digite - 02 - Para NÃO\n\n"
        "Temos uma informação importante referente a empresa %NomeCredor%\n\n"
        "Caso eu demora a responder, me chama no contato abaixo:\n\n"
        "Central de Atendimento e Negociações: wa.me://5591991600118"
    )
    tpl_quebra = (
        "Olá, tudo bem contigo ?\n\n"
        "Notificamos %Nome%, portador do documento %CpfCnpjMascarado%, que por motivo de não identificarmos os pagamentos das parcelas do acordo feito junto a empresa %NomeCredor%, o mesmo esta sendo cancelado e encaminhado para inclusão em protesto cartorial.\n\n"
        "Em segunda estancia, será levado ao acionamento juridico no Forum da comarca de sua residencia.\n\n"
        "A divida hoje encontra-se no valor de *R$ %ValorTotalParcelas%*H\n\n"
        "Fico no aguardo do seu contato nas proxima 24 horas para evitar o procedimento acima informado.\n\n"
        "NOSSO CANAL DE ATENDIMENTO\n"
        "(91) 99160-0118\n"
        "®NEGOCIAR COBRANÇA\n\n\n"
        "atenciosamente\n"
        "%NomeConsultor%"
    )

    base_data = {
        "%Nome%": (devedor.nome or "").strip(),
        "%CpfCnpjMascarado%": cpf_cnpj_mascarado,
        "%NomeConsultor%": nome_consultor,
        "%NomeCredor%": nome_credor,
    }

    def _apply(tpl, mapping):
        for k, v in mapping.items():
            tpl = tpl.replace(k, v)
        return tpl

    data_vencidas = {
        **base_data,
        "%QtdeParcelas%": str(qtde_vencidas),
        "%ListaVencimentosParcelas%": (lista_vencidas or "-"),
        "%ValorTotalParcelas%": _format_brl(total_vencidas),   # COM JUROS
    }

    msg_vencidas = _apply(tpl_vencidas, data_vencidas)
    msg_a_vencer = _apply(tpl_a_vencer, base_data)
    msg_padrao   = _apply(tpl_padrao,   base_data)
    msg_quebra   = _apply(tpl_quebra,   {**base_data, "%ValorTotalParcelas%": _format_brl(total_quebra)})

    # ----- Óbito (opcional, nunca derruba a página)
    obito_info = {}
    if devedor.cpf:
        try:
            obito_info = consultar_obito(devedor.cpf)
            if obito_info.get("deceased"):
                messages.warning(request, "⚠️ Registro de óbito localizado para este CPF.")
        except Exception as e:
            obito_info = {"checked": False, "deceased": False, "status": "ERROR", "error": str(e)}

    # ----- POST: salvar telefones
    if request.method == "POST":
        telefone_fields = [f"telefone{i}" for i in range(1, 11)]
        valido_fields = [f"telefone{i}_valido" for i in range(1, 11)]
        try:
            for f in telefone_fields:
                v = (request.POST.get(f, "") or "").strip()
                if v:
                    setattr(devedor, f, v)
            for f in valido_fields:
                v = (request.POST.get(f, "") or "").strip()
                if v:
                    setattr(devedor, f, v)
            devedor.save()
            messages.success(request, "Telefones salvos com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro ao salvar os telefones: {e}")
        return redirect("detalhes_devedor", titulo_id=titulo.id)

    context = {
        "titulo": titulo,
        "devedor": devedor,
        "empresa": getattr(devedor, "empresa", None),
        "titulos": titulos,
        "titulos_entrada": list(titulos_entrada),
        "titulos_associados": list(titulos_associados),
        "agendamentos": Agendamento.objects.filter(devedor=devedor),
        "follow_ups": follow_ups,

        # mensagens Whats
        "msg_vencidas": msg_vencidas,
        "msg_a_vencer": msg_a_vencer,
        "msg_padrao": msg_padrao,
        "msg_quebra": msg_quebra,
        "mensagem_whatsapp": msg_padrao,

        # totais/proposta
        "total_quitado": total_quitado,
        "total_negociado": total_negociado,
        "total_pendente": total_pendente_valor,
        "saldo_pendente": float(saldo_pendente),

        "percentual_avista_default": 45,
        "percentual_parc_default": 55,

        "forma_pagamento_map": forma_pagamento_map,
        "today": hoje,
        "obito_info": obito_info,
    }
    return render(request, "detalhes_devedor.html", context)

def consult_api(request):
    if request.method == "POST":
        data = json.loads(request.body)
        devedores_ids = data.get("devedores", [])

        access_token = "nG1dFp8huYPdgkhpWbAvg0TPYMRLx90B41eOLaoq"  # Replace with your actual access token

        for devedor_id in devedores_ids:
            devedor = get_object_or_404(Devedor, id=devedor_id)

            # Check if CPF or CNPJ is available
            cpf = re.sub(r"\D", "", devedor.cpf) if devedor.cpf else None
            cnpj = re.sub(r"\D", "", devedor.cnpj) if devedor.cnpj else None

            # Determine the correct endpoint based on CPF or CNPJ
            if cpf:
                url = f"https://api.lemit.com.br/api/v1/consulta/pessoa/{cpf}"
            elif cnpj:
                url = f"https://api.lemit.com.br/api/v1/consulta/empresa/{cnpj}"
            else:
                return JsonResponse(
                    {"success": False, "message": "Neither CPF nor CNPJ is available"}
                )

            headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            }

            try:
                response = requests.get(url, headers=headers)
                response.raise_for_status()

                api_data = response.json()

                if "pessoa" in api_data:
                    pessoa_data = api_data["pessoa"]
                elif "empresa" in api_data:
                    pessoa_data = api_data["empresa"]
                else:
                    return JsonResponse(
                        {"success": False, "message": "Invalid API response structure"}
                    )

                devedor.nome = pessoa_data.get("nome", devedor.nome)
                devedor.nome_mae = pessoa_data.get("nome_mae", devedor.nome_mae)

                celulares = pessoa_data.get("celulares", [])
                fixos = pessoa_data.get("fixos", [])

                # Update phone fields, filling empty fields first
                phones = [f"{c['ddd']}{c['numero']}" for c in celulares] + [
                    f"{f['ddd']}{f['numero']}" for f in fixos
                ]

                phone_fields = [
                    "telefone1",
                    "telefone2",
                    "telefone3",
                    "telefone4",
                    "telefone5",
                    "telefone6",
                    "telefone7",
                    "telefone8",
                    "telefone9",
                    "telefone10",
                ]

                # Fill empty phone fields first
                empty_fields = [
                    field
                    for field in phone_fields
                    if getattr(devedor, field) in [None, ""]
                ]
                for phone in phones:
                    if empty_fields:
                        setattr(devedor, empty_fields.pop(0), phone)
                    else:
                        # If no empty fields, replace existing fields one by one
                        for field in phone_fields:
                            setattr(devedor, field, phone)
                            phone_fields.remove(field)
                            break

                devedor.save()

            except requests.exceptions.HTTPError as http_err:
                return JsonResponse(
                    {
                        "success": False,
                        "message": f"HTTP error occurred: {http_err}. Response: {response.content.decode()}",
                    }
                )
            except Exception as err:
                return JsonResponse(
                    {"success": False, "message": f"Other error occurred: {err}"}
                )

        return JsonResponse({"success": True})
    return JsonResponse({"success": False, "message": "Invalid request method"})

from django.db.models import Q
from django.utils.timezone import now
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required

@login_required
def proximo_cliente(request, titulo_id):
    hoje = now().date()
    # título/devedor atual
    t_atual = get_object_or_404(Titulo.objects.select_related('devedor__empresa'), pk=titulo_id)
    dev_atual_id = t_atual.devedor_id

    # Pool de títulos elegíveis
    qs = Titulo.objects.filter(
        Q(statusBaixa__in=[0, 3]) | Q(statusBaixa__isnull=True)  # pendentes/negociados
    )

    # (opcional) limitar ao operador logado OU sem operador
    qs = qs.filter(Q(operador=request.user.username) | Q(operador__isnull=True) | Q(operador=""))

    # Excluir só os que já tiveram ação HOJE (em vez de exigir isnull)
    qs = qs.exclude(ultima_acao=hoje)

    # Empresas ativas (via devedor -> empresa)
    qs = qs.filter(devedor__empresa__status_empresa=True)

    # Devedores únicos em ordem determinística
    devedores_ids = list(qs.order_by('devedor_id').values_list('devedor_id', flat=True).distinct())
    if not devedores_ids:
        messages.info(request, 'Nenhum devedor elegível encontrado.')
        return redirect('dashboard')

    try:
        idx = devedores_ids.index(dev_atual_id)
    except ValueError:
        idx = -1

    prox_dev_id = devedores_ids[(idx + 1) % len(devedores_ids)]

    prox_titulo = (
        Titulo.objects.filter(devedor_id=prox_dev_id)
        .filter(Q(statusBaixa__in=[0, 3]) | Q(statusBaixa__isnull=True))
        .order_by('dataVencimento', 'id')
        .first()
    ) or Titulo.objects.filter(devedor_id=prox_dev_id).order_by('id').first()

    if prox_titulo:
        return redirect('detalhes_devedor', titulo_id=prox_titulo.id)

    return redirect('listar_devedores')


@login_required
def lista_titulos(request):
    titulos = Titulo.objects.all()  # Ajuste a lógica conforme necessário
    return render(request, "lista_titulos.html", {"titulos": titulos})


@login_required
def editar_telefones(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)
    if request.method == "POST":
        devedor.telefone1 = request.POST.get("telefone1")
        devedor.telefone2 = request.POST.get("telefone2")
        devedor.telefone3 = request.POST.get("telefone3")
        devedor.telefone4 = request.POST.get("telefone4")
        devedor.telefone5 = request.POST.get("telefone5")
        devedor.telefone6 = request.POST.get("telefone6")
        devedor.telefone7 = request.POST.get("telefone7")
        devedor.telefone8 = request.POST.get("telefone8")
        devedor.telefone9 = request.POST.get("telefone9")
        devedor.telefone10 = request.POST.get("telefone10")
        devedor.save()
        # messages.success(request, "Telefones atualizados com sucesso!")
        return redirect("detalhes_devedor", titulo_id=devedor.id)
    return redirect("detalhes_devedor", titulo_id=devedor.id)


@login_required
def editar_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    devedores = Devedor.objects.all()
    empresas = Empresa.objects.all()  # Fetching all core_empresa
    if request.method == "POST":
        data = request.POST
        agendamento.devedor = Devedor.objects.get(id=data["devedor_id"])
        agendamento.empresa = Empresa.objects.get(id=data["empresa_id"])
        agendamento.data_abertura = data["data_abertura"]
        agendamento.data_retorno = data["data_retorno"]
        agendamento.assunto = data["assunto"]
        agendamento.operador = data.get("operador", "")
        agendamento.save()
        # messages.success(request, 'Agendamento atualizado com sucesso.')
        return redirect("listar_agendamentos")
    return render(
        request,
        "agendamentos_editar.html",
        {
            "agendamento": agendamento,
            "devedores": devedores,
            "empresas": empresas,  # Use 'empresas' here instead of 'core_empresa'
        },
    )


# Excluir Agendamento
@login_required
def excluir_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    agendamento.delete()
    # messages.success(request, 'Agendamento excluído com sucesso.')
    return redirect("listar_agendamentos")


@login_required
def listar_follow_ups(request, devedor_id):
    devedor = get_object_or_404(Devedor, id=devedor_id)
    follow_ups = FollowUp.objects.filter(devedor=devedor).order_by("-created_at")
    return render(
        request,
        "follow_ups_listar.html",
        {"devedor": devedor, "follow_ups": follow_ups},
    )


@login_required
def adicionar_follow_up(request, devedor_id):
    if request.method == "POST":
        devedor = get_object_or_404(Devedor, id=devedor_id)
        texto = request.POST.get("texto")

        if texto:
            FollowUp.objects.create(devedor=devedor, texto=texto)
            # messages.success(request, "Follow-up adicionado com sucesso.")
        else:
            messages.error(request, "O texto do Follow-up não pode estar vazio.")

        # Tentar encontrar um título relacionado
        titulo = Titulo.objects.filter(devedor=devedor).first()
        if titulo:
            return redirect("detalhes_devedor", titulo_id=titulo.id)
        else:
            messages.warning(request, "Nenhum título encontrado para o devedor.")
            return redirect("lista_devedores")  # Substitua por uma view apropriada
    else:
        messages.error(request, "Método inválido.")
        return redirect("lista_devedores")


def listar_logs(request):
    # Obtém todos os logs de acesso
    logs = UserAccessLog.objects.all().order_by("-timestamp")

    # Cria o paginador
    paginator = Paginator(logs, 30)  # 30 logs por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Passa os logs paginados para o template
    return render(request, "logs_listar.html", {"page_obj": page_obj})


def buscar_devedores(request):
    if request.method == "GET":
        termo = request.GET.get("termo", "").strip()
        if termo:
            devedores = Devedor.objects.filter(nome__icontains=termo)[
                :10
            ]  # Limitar os resultados a 10
            resultados = [
                {
                    "id": devedor.id,
                    "nome": devedor.nome,
                    "empresa_nome": (
                        devedor.empresa.nome_fantasia
                        if devedor.empresa
                        else "Não associado"
                    ),
                }
                for devedor in devedores
            ]
            return JsonResponse(resultados, safe=False)
        return JsonResponse([], safe=False)


def configurar_permissoes_admin():

    # Obter ou criar o grupo Admin
    admin_group, _ = Group.objects.get_or_create(name="Admin")

    # Associar todas as permissões disponíveis ao grupo Admin
    todas_permissoes = Permission.objects.all()
    admin_group.permissions.set(todas_permissoes)
    admin_group.save()

    print(f"O grupo '{admin_group.name}' agora tem todas as permissões.")


def listar_mensagens(request):
    mensagens = MensagemWhatsapp.objects.all()
    return render(request, "mensagens_listar.html", {"mensagens": mensagens})


def adicionar_mensagem(request):
    if request.method == "POST":
        form = MensagemWhatsappForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("listar_mensagens")
    else:
        form = MensagemWhatsappForm()
    return render(request, "mensagem_adicionar.html", {"form": form})


def editar_mensagem(request, pk):
    mensagem = get_object_or_404(MensagemWhatsapp, pk=pk)
    if request.method == "POST":
        form = MensagemWhatsappForm(request.POST, instance=mensagem)
        if form.is_valid():
            form.save()
            return redirect("listar_mensagens")
    else:
        form = MensagemWhatsappForm(instance=mensagem)
    return render(request, "mensagem_editar.html", {"form": form})


def excluir_mensagem(request, pk):
    mensagem = get_object_or_404(MensagemWhatsapp, pk=pk)
    mensagem.delete()
    return redirect("listar_mensagens")


def tabelas_listar(request):
    query = request.GET.get("q", "")
    tabelas = (
        TabelaRemuneracao.objects.filter(nome__icontains=query)
        if query
        else TabelaRemuneracao.objects.all()
    )

    paginator = Paginator(tabelas, 10)  # 10 tabelas por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request, "tabelas_listar.html", {"page_obj": page_obj, "query": query}
    )


def tabela_adicionar(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        if nome:
            TabelaRemuneracao.objects.create(nome=nome)
            # messages.success(request, "Tabela adicionada com sucesso!")
        return redirect("tabelas_listar")
    return render(request, "tabela_adicionar.html")


def tabela_editar(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    if request.method == "POST":
        tabela.nome = request.POST.get("nome")
        tabela.save()
        # messages.success(request, "Tabela editada com sucesso!")
        return redirect("tabelas_listar")
    return render(request, "tabela_editar.html", {"tabela": tabela})


def tabela_excluir(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    tabela.delete()
    # messages.success(request, "Tabela excluída com sucesso!")
    return redirect("tabelas_listar")


def lista_gerenciar(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)

    if request.method == "POST":
        # Coleta os dados do formulário
        de_dias = request.POST.get("de_dias")
        ate_dias = request.POST.get("ate_dias")
        percentual_remuneracao = request.POST.get("percentual_remuneracao")

        # Criação de um novo item na tabela
        TabelaRemuneracaoLista.objects.create(
            tabela_remuneracao=tabela,
            de_dias=de_dias,
            ate_dias=ate_dias,
            percentual_remuneracao=percentual_remuneracao,
        )

        # Mensagem de sucesso
        messages.success(request, "Item adicionado à lista!")

    # Obtém todos os itens relacionados à tabela
    itens = tabela.listas.all()

    # Renderiza o template com o nome da tabela e os itens
    return render(request, "lista_gerenciar.html", {"tabela": tabela, "itens": itens})


def lista_editar(request, tabela_id, item_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    item = get_object_or_404(
        TabelaRemuneracaoLista, id=item_id, tabela_remuneracao=tabela
    )

    if request.method == "POST":
        item.de_dias = request.POST.get("de_dias")
        item.ate_dias = request.POST.get("ate_dias")

        # Substituir vírgula por ponto no percentual_remuneracao
        percentual_remuneracao = request.POST.get("percentual_remuneracao", "").replace(
            ",", "."
        )

        try:
            item.percentual_remuneracao = Decimal(percentual_remuneracao)
            item.save()
            # messages.success(request, "Item atualizado com sucesso!")
            return redirect("lista_gerenciar", tabela_id=tabela.id)
        except Exception as e:
            messages.error(request, f"Erro ao atualizar item: {e}")

    return render(request, "lista_editar.html", {"tabela": tabela, "item": item})


def lista_adicionar(request, tabela_id):
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    if request.method == "POST":
        de_dias = request.POST.get("de_dias")
        ate_dias = request.POST.get("ate_dias")
        percentual_remuneracao = request.POST.get("percentual_remuneracao")
        if de_dias and ate_dias and percentual_remuneracao:
            TabelaRemuneracaoLista.objects.create(
                tabela_remuneracao=tabela,
                de_dias=de_dias,
                ate_dias=ate_dias,
                percentual_remuneracao=percentual_remuneracao,
            )
            # messages.success(request, "Item adicionado à lista com sucesso!")
        else:
            messages.error(request, "Todos os campos são obrigatórios!")
        return redirect("lista_gerenciar", tabela_id=tabela.id)
    return render(request, "lista_adicionar.html", {"tabela": tabela})


def lista_excluir(request, lista_id):
    lista = get_object_or_404(TabelaRemuneracaoLista, id=lista_id)
    tabela_id = (
        lista.tabela_remuneracao.id
    )  # Captura o ID da tabela para redirecionar após exclusão
    lista.delete()
    # messages.success(request, "Item da lista excluído com sucesso!")
    return redirect("lista_gerenciar", tabela_id=tabela_id)


@login_required
def quitados_listar(request):
    # Obter os parâmetros de filtro de data
    data_inicio = request.GET.get("data_inicio", "")
    data_fim = request.GET.get("data_fim", "")

    # Construir a consulta base
    query = """
        SELECT 
            titulo.data_baixa,
            titulo.dataVencimento,
            titulo.valorRecebido,
            devedores.nome,
            devedores.cpf,
            devedores.cnpj,
            core_empresa.nome_fantasia,
            titulo.idTituloRef
        FROM 
            titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE titulo.data_baixa is not null and titulo.valorRecebido is not null
    """

    # Filtros adicionais de data
    params = []
    if data_inicio:
        query += " AND titulo.data_baixa >= %s"
        params.append(data_inicio)
    if data_fim:
        query += " AND titulo.data_baixa <= %s"
        params.append(data_fim)

    # Ordenar os resultados
    query += " ORDER BY titulo.data_baixa DESC;"

    # Executar a consulta
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    # Converter os resultados
    quitados = [
        {
            "data_baixa": row[0].strftime("%d/%m/%Y") if row[0] else "",
            "data_vencimento": row[1].strftime("%d/%m/%Y") if row[1] else "",
            "valor_recebido": row[2] or 0.0,  # Tratar None como 0.0
            "nome": row[3],
            "cpf": row[4],
            "cnpj": row[5],
            "empresa": row[6],
            "idTituloRef": row[7],
        }
        for row in rows
    ]

    # Paginação
    paginator = Paginator(quitados, 10)  # 10 itens por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Calcular a soma dos valores recebidos
    soma_total = sum(item["valor_recebido"] for item in quitados)

    return render(
        request,
        "quitados.html",
        {
            "page_obj": page_obj,
            "soma_total": soma_total,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
        },
    )


@login_required
def anexar_contrato(request, titulo_id):
    titulo = get_object_or_404(Titulo, pk=titulo_id)
    if request.method == "POST" and "contrato" in request.FILES:
        contrato_file = request.FILES["contrato"]
        extension = os.path.splitext(contrato_file.name)[1]
        unique_filename = f"{uuid.uuid4()}{extension}"
        titulo.contrato.save(unique_filename, contrato_file)
        titulo.save()
        # messages.success(request, "Contrato anexado com sucesso!")
        return redirect("acordos_listar")
    else:
        messages.error(request, "Falha ao anexar o contrato. Tente novamente.")
        return redirect("acordos_listar")


@login_required
def baixar_contrato_view(request, titulo_id):
    titulo = get_object_or_404(Titulo, pk=titulo_id)
    contrato_path = titulo.contrato.path
    with open(contrato_path, "rb") as fh:
        response = HttpResponse(fh.read(), content_type="application/force-download")
        response["Content-Disposition"] = (
            f'attachment; filename="{os.path.basename(contrato_path)}"'
        )
        return response


@login_required
def usuarios_lojista_listar(request):
    query = request.GET.get("q", "")  # Obtem o valor da pesquisa
    usuarios = UsersLojistas.objects.filter(
        Q(name__icontains=query)
        | Q(email__icontains=query)
        | Q(empresa__razao_social__icontains=query)
    )

    # Paginação
    paginator = Paginator(usuarios, 10)  # Mostra 10 usuários por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request, "usuarios_lojista_listar.html", {"page_obj": page_obj, "query": query}
    )


@login_required
def usuarios_lojista_criar(request):
    if request.method == "POST":
        nome = request.POST.get("name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        empresa_id = request.POST.get("empresa_id")

        if UsersLojistas.objects.filter(email=email).exists():
            messages.error(request, "Email já cadastrado.")
        else:
            try:
                empresa = Empresa.objects.get(id=empresa_id)

                # Hash da senha com bcrypt
                hashed_password = bcrypt.hashpw(
                    password.encode("utf-8"), bcrypt.gensalt()
                ).decode("utf-8")
                hashed_password = hashed_password.replace("$2b$", "$2y$", 1)

                lojista = UsersLojistas(
                    name=nome, email=email, password=hashed_password, empresa=empresa
                )
                lojista.save()
                # messages.success(request, "Usuário criado com sucesso.")
                return redirect("usuarios_lojista_listar")
            except Empresa.DoesNotExist:
                messages.error(request, "Empresa inválida.")

    empresas = Empresa.objects.all()
    return render(request, "usuarios_lojista_criar.html", {"empresas": empresas})


@login_required
def usuarios_lojista_editar(request, user_id):
    user = get_object_or_404(UsersLojistas, id=user_id)

    if request.method == "POST":
        user.name = request.POST.get("name")
        user.email = request.POST.get("email")
        new_password = request.POST.get("password")

        if new_password:
            # Hash da nova senha
            hashed_password = bcrypt.hashpw(
                new_password.encode("utf-8"), bcrypt.gensalt()
            ).decode("utf-8")
            hashed_password = hashed_password.replace("$2b$", "$2y$", 1)
            user.password = hashed_password

        user.save()
        # messages.success(request, "Usuário atualizado com sucesso.")
        return redirect("usuarios_lojista_listar")

    empresas = Empresa.objects.all()
    return render(
        request, "usuarios_lojista_editar.html", {"user": user, "empresas": empresas}
    )


@login_required
def emails_envio_criar(request):
    if request.method == "POST":
        email = request.POST.get("email")
        autenticacao = request.POST.get("autenticacao")
        porta = request.POST.get("porta")
        servidor_smtp = request.POST.get("servidor_smtp")
        tipo_envio = request.POST.get("tipo_envio")
        provedor = request.POST.get("provedor")
        senha = request.POST.get("senha")

        EmailEnvio.objects.create(
            email=email,
            autenticacao=autenticacao,
            porta=porta,
            servidor_smtp=servidor_smtp,
            tipo_envio=tipo_envio,
            provedor=provedor,
            senha=senha,
        )

        messages.success(request, "E-mail de envio cadastrado com sucesso.")
        return redirect("emails_envio_listar")

    return render(request, "emails_envio_criar.php")


@login_required
def emails_envio_editar(request, id):
    email_envio = get_object_or_404(EmailEnvio, id=id)

    if request.method == "POST":
        email_envio.email = request.POST.get("email")
        email_envio.autenticacao = request.POST.get("autenticacao")
        email_envio.porta = request.POST.get("porta")
        email_envio.servidor_smtp = request.POST.get("servidor_smtp")
        email_envio.tipo_envio = request.POST.get("tipo_envio")
        email_envio.provedor = request.POST.get("provedor")
        email_envio.senha = request.POST.get("senha")
        email_envio.save()

        messages.success(request, "E-mail de envio atualizado com sucesso.")
        return redirect("emails_envio_listar")

    return render(request, "emails_envio_editar.php", {"email_envio": email_envio})


@login_required
def emails_envio_listar(request):
    emails = EmailEnvio.objects.all()
    return render(request, "emails_envio_listar.php", {"emails": emails})


@login_required
def email_template_listar(request):
    templates = EmailTemplate.objects.all()
    return render(request, "email_template_listar.html", {"templates": templates})


@login_required
def email_template_criar(request):
    if request.method == "POST":
        tipo_envio = request.POST.get("tipo_envio")
        mensagem = request.POST.get("mensagem")
        EmailTemplate.objects.create(tipo_envio=tipo_envio, mensagem=mensagem)
        messages.success(request, "Template de e-mail criado com sucesso.")
        return redirect("email_template_listar")
    return render(request, "email_template_criar.html")


@login_required
def email_template_editar(request, id):
    template = get_object_or_404(EmailTemplate, id=id)
    if request.method == "POST":
        template.tipo_envio = request.POST.get("tipo_envio")
        template.mensagem = request.POST.get("mensagem")
        template.save()
        messages.success(request, "Template de e-mail atualizado com sucesso.")
        return redirect("email_template_listar")
    return render(request, "email_template_editar.html", {"template": template})


@login_required
@csrf_exempt
def emitir_boletos_view(request):
    logger.debug("Renderizando página de emissão de boletos...")

    # Consulta inicial para exibir os boletos disponíveis
    query = """
             SELECT 
                core_empresa.id AS empresa_id,
                core_empresa.razao_social,
                core_empresa.cnpj,
                core_empresa.endereco, 
                core_empresa.bairro,
                core_empresa.cidade,
                core_empresa.uf,
                core_empresa.cep,
                core_empresa.telefone, 
                GROUP_CONCAT(devedores.id ORDER BY devedores.id SEPARATOR ',') AS devedor_ids,
                MAX(DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira)) AS dias_de_atraso,
                SUM(titulo.valorRecebido) AS total_valorRecebido,
                SUM(
                    ROUND(
                        CASE 
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) BETWEEN 30 AND 45 THEN titulo.valorRecebido * 0.06
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) BETWEEN 46 AND 90 THEN titulo.valorRecebido * 0.09
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) BETWEEN 91 AND 120 THEN titulo.valorRecebido * 0.12
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) BETWEEN 121 AND 180 THEN titulo.valorRecebido * 0.15
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) BETWEEN 181 AND 360 THEN titulo.valorRecebido * 0.18
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) BETWEEN 361 AND 720 THEN titulo.valorRecebido * 0.23
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) BETWEEN 721 AND 1095 THEN titulo.valorRecebido * 0.26
                            WHEN DATEDIFF(titulo.data_baixa, titulo.dataVencimentoPrimeira) > 1095 THEN titulo.valorRecebido * 0.30
                            ELSE 0
                        END, 2)
                ) AS total_comissao_valor
                
            FROM 
                titulo, 
                devedores, 
                core_empresa
            WHERE 
                titulo.devedor_id = devedores.id 
                AND devedores.empresa_id = core_empresa.id 
                AND titulo.id_cobranca IS NULL 
                AND titulo.statusBaixa = 2 
                AND (email_enviado IS NULL OR email_enviado = "NAO")
            GROUP BY 
                core_empresa.id, 
                core_empresa.razao_social, 
                core_empresa.cnpj, 
                core_empresa.endereco, 
                core_empresa.bairro, 
                core_empresa.cidade, 
                core_empresa.uf, 
                core_empresa.cep, 
                core_empresa.telefone;

    """
    with connection.cursor() as cursor:
        cursor.execute(query)
        results = cursor.fetchall()

    headers = [
        "ID Empresa",
        "Razão Social",
        "CNPJ",
        "Endereço",
        "Bairro",
        "Cidade",
        "UF",
        "CEP",
        "Telefone",
        "ID Devedor(es)",
        "Dias de Atraso",
        "Valor Recebido",
        "Comissão (R$)",
    ]

    if request.method == "POST" and "emitir" in request.POST:
        try:
            # Capturar os dados enviados no botão "emitir"
            razao_social = request.POST.get("emitir")
            dias_atraso = request.POST.get("row10")
            valor_recebido = request.POST.get("row11")
            comissao = request.POST.get("row12")

            if not all([razao_social, dias_atraso, valor_recebido, comissao]):
                raise ValueError("Campos obrigatórios estão ausentes.")

            # Log de depuração para confirmar recebimento
            logger.debug(
                f"Dados recebidos no POST: razao_social={razao_social}, "
                f"dias_atraso={dias_atraso}, valor_recebido={valor_recebido}, comissao={comissao}"
            )

            # Verificar valores obrigatórios
            missing_fields = []
            if not razao_social:
                missing_fields.append("Razão Social")
            if not dias_atraso:
                missing_fields.append("Dias de Atraso")
            if not valor_recebido:
                missing_fields.append("Valor Recebido")
            if not comissao:
                missing_fields.append("Comissão")

            if missing_fields:
                raise ValueError(
                    f"Valores obrigatórios faltando: {', '.join(missing_fields)}."
                )

            # Buscar os dados necessários no banco
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT total_valorRecebido as valor_recebido,
                            total_comissao_valor as comissao_valor,
                            razao_social,
                            cnpj,
                            endereco,
                            bairro,
                            cidade,
                            uf,
                            cep,
                            data_vencimento,
                            devedor_ids
                    FROM boleto_temp
                    WHERE razao_social = %s AND dias_de_atraso = %s AND total_valorRecebido = %s AND total_comissao_valor = %s
                """,
                    [razao_social, dias_atraso, valor_recebido, comissao],
                )
                result = cursor.fetchone()

            if not result:
                raise Exception(
                    f"Nenhum registro encontrado para os critérios fornecidos: "
                    f"Razão Social={razao_social}, Dias de Atraso={dias_atraso}, "
                    f"Valor Recebido={valor_recebido}, Comissão={comissao}."
                )

            # Extraindo os dados corretamente
            (
                valor_recebido,
                comissao_valor,
                razao_social,
                cnpj,
                endereco,
                bairro,
                cidade,
                uf,
                cep,
                data_vencimento,
                devedor_ids,
            ) = result

            if not comissao_valor or comissao_valor <= 0:
                raise ValueError(f"O valor da comissão é inválido: {comissao_valor}")

            # Validar e separar os devedor_ids
            devedor_ids_list = [
                id.strip() for id in devedor_ids.split(",") if id.strip().isdigit()
            ]

            if not devedor_ids_list:
                raise ValueError("Nenhum ID de devedor válido encontrado na consulta.")

            # Emitir boleto completo
            resultado = emitir_boleto_inter_completo(
                valor=comissao_valor,
                nome_cliente=razao_social,
                pagador_cnpj=cnpj,
                beneficiario_cnpj="00082059000176",
                beneficiario_nome="M C A DA SILVA CONSULTORIA",
                beneficiario_endereco=f"{endereco}, {bairro}, {cidade}, {uf}, {cep}",
                data_vencimento=data_vencimento,
                endereco=endereco,
                bairro=bairro,
                cidade=cidade,
                uf=uf,
                cep=cep,
                devedor_id=None,  # Não necessário no momento da emissão
            )

            codigo_solicitacao = resultado.get("codigoSolicitacao")

            # Atualizar o campo id_cobranca para cada devedor_id
            with connection.cursor() as cursor:
                for devedor_id in devedor_ids_list:
                    cursor.execute(
                        "UPDATE titulo SET id_cobranca = %s WHERE devedor_id = %s",
                        [codigo_solicitacao, devedor_id],
                    )
                connection.commit()

            # Adicionar mensagem de sucesso
            messages.success(
                request,
                f"Boleto emitido com sucesso para os devedores {', '.join(devedor_ids_list)}! Código da cobrança: {codigo_solicitacao}",
            )

        except ValueError as ve:
            logger.error(f"Erro de validação: {ve}")
            messages.error(request, f"Erro de validação: {ve}")
        except Exception as e:
            logger.error(f"Erro ao emitir boleto: {e}")
            messages.error(request, f"Erro ao emitir boleto: {e}")

    context = {
        "headers": headers,
        "results": results,
    }

    return render(request, "boletos_emitir.html", context)


def baixar_boleto(request, codigo_solicitacao):
    """
    Faz o download do PDF do boleto com base no código de solicitação.
    """
    try:
        token = obter_token_inter()

        url = f"https://cdpj.partners.bancointer.com.br/cobranca/v3/cobrancas/{codigo_solicitacao}/pdf"

        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/pdf",
        }
        certificado = (settings.CERT_PATH, settings.KEY_PATH)

        response = requests.get(url, headers=headers, cert=certificado)
        response.raise_for_status()

        # Salvando em um arquivo temporário para download
        temp_path = f"/tmp/boleto_{codigo_solicitacao}.pdf"
        with open(temp_path, "wb") as temp_file:
            temp_file.write(response.content)

        # Retornar o arquivo para download
        return FileResponse(
            open(temp_path, "rb"),
            as_attachment=True,
            filename=f"boleto_{codigo_solicitacao}.pdf",
        )
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro ao baixar boleto para código {codigo_solicitacao}: {e}")
        return HttpResponse("Erro ao baixar boleto.", status=500)


def dictfetchall(cursor):
    """Converte os resultados da query em uma lista de dicionários."""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def format_date(value):
    """Garante que todas as datas sejam formatadas no padrão dd/mm/yyyy"""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")  # Se já for um datetime
    elif isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").strftime(
                "%d/%m/%Y"
            )  # Formato YYYY-MM-DD
        except ValueError:
            return value  # Se não conseguir converter, retorna o original
    return value


